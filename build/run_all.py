"""Run the full standalone V06 dashboard data pipeline."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parent.parent
BUILD_DIR = PROJECT_ROOT / "build"
WEB_DIR = PROJECT_ROOT / "web"
WEB_DATA_DIR = WEB_DIR / "public" / "data"
WORKBOOK_PATH = BUILD_DIR / "MuhuratFinder_V06_Workbook.xlsx"
RAW_BUILDER = BUILD_DIR / "dependencies" / "v06_workbook_builder.py"
PARENT_STATE_ENGINE = BUILD_DIR / "dependencies" / "v06_parent_state_engine.py"
EXPORTER = BUILD_DIR / "export_excel_to_json.py"
PUSH_SCRIPT = BUILD_DIR / "push_online.py"
REQUIRED_JSON_FILES = [
    WEB_DATA_DIR / "config.json",
    WEB_DATA_DIR / "day_summary.json",
    WEB_DATA_DIR / "muhurat-data.json",
    WEB_DATA_DIR / "windows.json",
]
EXTERNAL_PATH_PATTERNS = [
    re.compile(r"[\\\\/]v0[2-5][\\\\/]"),
    re.compile(r"\.\.[\\\\/]"),
    re.compile(r"\b[A-Za-z]:[\\\\/](?:[^\\\\/:*?\"<>|\r\n]+[\\\\/])+"),
]


def run_step(command: list[str], cwd: Path) -> None:
    print(f"RUN={' '.join(command)}")
    subprocess.run(command, cwd=cwd, check=True)


def validate_paths() -> None:
    for path in [BUILD_DIR, WEB_DIR, RAW_BUILDER, PARENT_STATE_ENGINE, EXPORTER, PUSH_SCRIPT]:
        if not path.exists():
            raise FileNotFoundError(f"Required V06 path missing: {path}")
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            f"Main workbook missing: {WORKBOOK_PATH}\n"
            "Restore it from version control or copy a known-good V06 workbook into build/."
        )


def validate_runtime_scripts_are_local() -> None:
    suspicious: list[str] = []
    for script_path in [RAW_BUILDER, PARENT_STATE_ENGINE, EXPORTER, PUSH_SCRIPT]:
        content = script_path.read_text(encoding="utf-8")
        for pattern in EXTERNAL_PATH_PATTERNS:
            if pattern.search(content):
                suspicious.append(f"{script_path.name}: matches '{pattern.pattern}'")
    if suspicious:
        raise ValueError(
            "Found non-local runtime path references in V06 build scripts:\n"
            + "\n".join(suspicious)
        )


def validate_workbook_structure() -> None:
    from openpyxl import load_workbook

    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    try:
        required = {"CONFIG", "EPHEMERIS_RAW"}
        missing = required.difference(wb.sheetnames)
        if missing:
            raise ValueError(f"Workbook missing required sheets: {', '.join(sorted(missing))}")
    finally:
        wb.close()


def validate_json_outputs() -> None:
    missing = [path for path in REQUIRED_JSON_FILES if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing generated JSON files:\n" + "\n".join(str(path) for path in missing))
    for path in REQUIRED_JSON_FILES:
        json.loads(path.read_text(encoding="utf-8"))


def maybe_build_web(skip_web_build: bool) -> str:
    if skip_web_build:
        return "skipped by flag"
    npm = shutil.which("npm.cmd") or shutil.which("npm")
    if npm is None:
        return "skipped: npm not found"
    run_step([npm, "run", "build"], WEB_DIR)
    return "built"


def maybe_push(push: bool, message: str) -> str:
    if not push:
        return "not requested"
    run_step([sys.executable, str(PUSH_SCRIPT), "--message", message], PROJECT_ROOT)
    return "pushed"


def main() -> None:
    parser = argparse.ArgumentParser(description="Run the standalone V06 workbook-to-dashboard pipeline.")
    parser.add_argument("--push", action="store_true", help="Commit and push V06 after successful build.")
    parser.add_argument(
        "--message",
        default="Update dashboard data",
        help="Commit message used when --push is supplied.",
    )
    parser.add_argument(
        "--skip-web-build",
        action="store_true",
        help="Skip npm build. JSON export will still run.",
    )
    parser.add_argument(
        "--no-recalculate",
        action="store_true",
        help="Skip Excel COM recalculation during JSON export.",
    )
    args = parser.parse_args()

    validate_paths()
    validate_runtime_scripts_are_local()
    validate_workbook_structure()
    WEB_DATA_DIR.mkdir(parents=True, exist_ok=True)

    run_step([sys.executable, str(RAW_BUILDER)], PROJECT_ROOT)
    run_step([sys.executable, str(PARENT_STATE_ENGINE)], PROJECT_ROOT)

    export_command = [sys.executable, str(EXPORTER)]
    if args.no_recalculate:
        export_command.append("--no-recalculate")
    run_step(export_command, PROJECT_ROOT)

    validate_workbook_structure()
    validate_json_outputs()
    web_build_result = maybe_build_web(args.skip_web_build)
    push_result = maybe_push(args.push, args.message)

    print(f"WORKBOOK={WORKBOOK_PATH}")
    print(f"WEB_BUILD={web_build_result}")
    print(f"PUSH={push_result}")
    for path in REQUIRED_JSON_FILES:
        print(f"JSON={path}")


if __name__ == "__main__":
    main()
