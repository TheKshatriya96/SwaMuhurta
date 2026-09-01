"""Build the V06 standalone workbook base.

V06 workbook base only:
- raw ephemeris + Panchang
- event location
- natal reference
- helper/reference columns

No scoring columns.
No parent-state logic.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
import sys
from typing import Any

import pandas as pd
SCRIPT_DIR = Path(__file__).resolve().parent
BUILD_DIR = SCRIPT_DIR.parent
RECOMMENDED_PYTHON = BUILD_DIR / ".venv" / "Scripts" / "python.exe"
try:
    import swisseph as swe
except ModuleNotFoundError as exc:  # pragma: no cover - environment-specific guard
    raise SystemExit(
        "Missing dependency: swisseph.\n"
        "Install the build requirements or run this script with the dashboard virtualenv interpreter:\n"
        f"{RECOMMENDED_PYTHON} {SCRIPT_DIR / 'v06_workbook_builder.py'}"
    ) from exc
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_WORKBOOK = BUILD_DIR / "MuhuratFinder_V06_Workbook.xlsx"

if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from muhurta_engine.astronomy import SwissEphemerisEngine, local_datetime_to_jd  # noqa: E402
from muhurta_engine.config import DEFAULT_LOCATION, LocationConfig  # noqa: E402
from muhurta_engine.v03_raw_export import (  # noqa: E402
    RAW_HEADERS,
    build_day_payload,
    default_raw_config,
    find_segment,
    floor_minute,
    karana_name,
    longitude_parts,
    overlaps,
    tithi_details,
    whole_sign_house,
    write_sheet,
    yoga_name,
)


CONFIG_META: dict[str, str] = {}

CONFIG_CELL_MAP = {
    "from_date": "B4",
    "end_date": "B5",
    "ayanamsa": "B6",
    "default_timezone": "B7",
    "default_dst": "B8",
    "event_location_name": "E4",
    "event_latitude": "E5",
    "event_longitude": "E6",
    "event_timezone": "E7",
    "event_dst": "E8",
    "user_name": "B11",
    "natal_date": "B12",
    "natal_time": "B13",
    "natal_birth_place": "B14",
    "natal_latitude": "B15",
    "natal_longitude": "B16",
    "natal_timezone": "B17",
    "natal_moon_sign": "B18",
    "natal_nakshatra": "B19",
    "natal_lagna": "B20",
}

OLD_FORMULA_COLUMNS = [
    "TithiScore",
    "NakshatraScore",
    "YogaScore",
    "KaranaScore",
    "PakshaScore",
    "ChoghadiyaScore",
    "HoraScore",
    "AbhijitScore",
    "RahuPenalty",
    "YamagandaPenalty",
    "GulikaPenalty",
    "DurmuhurtaPenalty",
    "VarjyamPenalty",
    "TithiKshayaPenalty",
    "NakshatraKshayaPenalty",
    "LagnaDegreeScore",
    "LagnaSignScore",
    "LagnaNakshatraScore",
    "MaleficPressureScore",
    "MoonHouseScore",
    "MoonDegreeScore",
    "MoonNakshatraQualityScore",
    "VenusStrengthScore",
    "JupiterSupportScore",
    "MercuryCommunicationScore",
    "ChandrabalamScore",
    "TarabalamScore",
    "PanchangScore",
    "TimeFilterScore",
    "LagnaStrengthScore",
    "MoonStrengthScore",
    "PlanetarySupportScore",
    "PersonalCompatibilityScore",
    "OverallMuhuratScore",
    "RiskLevel",
    "WindowType",
    "BestFor",
    "AvoidFor",
    "WarningText",
    "MarriageContactScore",
    "BusinessMoneyScore",
    "TechnicalWorkScore",
    "SpiritualScore",
    "PrimaryState",
    "PrimaryStateReason",
    "SecondaryStates",
    "SecondaryStateReason",
    "BestActions",
    "AvoidActions",
]

EVENT_COLUMNS = [
    "EventLocationName",
    "EventLatitude",
    "EventLongitude",
    "EventTimezone",
    "EventDST",
]

NATAL_COLUMNS = [
    "NatalBirthPlace",
    "NatalLatitude",
    "NatalLongitude",
    "NatalTimezone",
    "NatalMoonSign",
    "NatalNakshatra",
    "NatalLagna",
]

TARA_CHANDRA_COLUMNS = [
    "CurrentMoonNakshatraNo",
    "ChandrabalamSignDistance",
    "TarabalamNakshatraDistance",
    "TaraCycleNo",
    "TaraName",
]

HOUSE_COLUMNS = [
    "LagnaLord",
    "House2Sign",
    "House2Lord",
    "House3Sign",
    "House3Lord",
    "House4Sign",
    "House4Lord",
    "House5Sign",
    "House5Lord",
    "House6Sign",
    "House6Lord",
    "House7Sign",
    "House7Lord",
    "House8Sign",
    "House8Lord",
    "House9Sign",
    "House9Lord",
    "House10Sign",
    "House10Lord",
    "House11Sign",
    "House11Lord",
    "House12Sign",
    "House12Lord",
]

DIGNITY_COLUMNS = [
    "SunDignity",
    "MoonDignity",
    "MarsDignity",
    "MercuryDignity",
    "JupiterDignity",
    "VenusDignity",
    "SaturnDignity",
]

NOTE_COLUMNS = [
    "LocationValidityNote",
    "CalculationLayerNote",
]

FOUNDATION_COLUMNS = EVENT_COLUMNS + NATAL_COLUMNS + TARA_CHANDRA_COLUMNS + HOUSE_COLUMNS + DIGNITY_COLUMNS + NOTE_COLUMNS


@dataclass(frozen=True)
class FoundationConfig:
    """Workbook config for the V06 standalone workbook base.

    Birth location affects natal chart.
    Event location affects muhurat.
    Do not calculate muhurat using natal birth location unless the person is physically there during the event.
    """

    from_date: date
    end_date: date
    ayanamsa: str
    default_timezone: str
    default_dst: str
    event_location_name: str
    event_latitude: float | None
    event_longitude: float | None
    event_timezone: str
    event_dst: str
    user_name: str
    natal_date: str
    natal_time: str
    natal_birth_place: str
    natal_latitude: str
    natal_longitude: str
    natal_timezone: str
    natal_moon_sign: str
    natal_nakshatra: str
    natal_lagna: str


def excel_col(col_index: int) -> str:
    return get_column_letter(col_index)


def safe_col(header_map: dict[str, int], header_name: str) -> int:
    if header_name not in header_map:
        raise KeyError(f"Header not found: {header_name}")
    return header_map[header_name]


def get_or_create_sheet(wb: Workbook, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


def clear_sheet(ws) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def get_header_map(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value not in (None, "")
    }


def default_foundation_config() -> FoundationConfig:
    base = default_raw_config()
    return FoundationConfig(
        from_date=base.from_date,
        end_date=base.end_date,
        ayanamsa=base.ayanamsa,
        default_timezone="Asia/Kolkata",
        default_dst="N",
        event_location_name="Mumbai, Maharashtra",
        event_latitude=DEFAULT_LOCATION.latitude,
        event_longitude=DEFAULT_LOCATION.longitude,
        event_timezone="Asia/Kolkata",
        event_dst="N",
        user_name="Swapnil",
        natal_date="22-Dec-1994",
        natal_time="23:05",
        natal_birth_place="Mahad, Maharashtra",
        natal_latitude="",
        natal_longitude="",
        natal_timezone="Asia/Kolkata",
        natal_moon_sign="Cancer",
        natal_nakshatra="Ashlesha",
        natal_lagna="Leo",
    )


def _as_date(value: Any, fallback: date) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d-%b-%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return fallback


def read_foundation_config(path: Path) -> FoundationConfig:
    wb = load_workbook(path, data_only=True)
    ws = wb["CONFIG"]
    defaults = default_foundation_config()
    legacy_map = {
        "from_date": "B3",
        "end_date": "B4",
        "ayanamsa": "B5",
        "default_timezone": "B6",
        "default_dst": "B7",
        "event_location_name": "E3",
        "event_latitude": "E4",
        "event_longitude": "E5",
        "event_timezone": "E6",
        "event_dst": "E7",
        "user_name": "B10",
        "natal_date": "B11",
        "natal_time": "B12",
        "natal_birth_place": "B13",
        "natal_latitude": "B14",
        "natal_longitude": "B15",
        "natal_timezone": "B16",
        "natal_moon_sign": "B17",
        "natal_nakshatra": "B18",
        "natal_lagna": "B19",
    }
    legacy_layout = ws["B3"].value not in (None, "") and ws["B10"].value not in (None, "") and ws["B20"].value in (None, "")

    def cell_value(key: str):
        if legacy_layout and key in legacy_map:
            return ws[legacy_map[key]].value
        value = ws[CONFIG_CELL_MAP[key]].value
        if value in (None, "") and key in legacy_map:
            return ws[legacy_map[key]].value
        if key in ("from_date", "end_date") and not isinstance(value, (date, datetime)):
            legacy_value = ws[legacy_map[key]].value
            if isinstance(legacy_value, (date, datetime)):
                return legacy_value
        return value

    def text(key: str, fallback: str = "") -> str:
        value = cell_value(key)
        if value is None:
            return fallback
        return str(value).strip()

    def number(key: str, fallback=None):
        value = cell_value(key)
        if value in (None, ""):
            return fallback
        try:
            return float(value)
        except (TypeError, ValueError):
            return fallback

    return FoundationConfig(
        from_date=_as_date(cell_value("from_date"), defaults.from_date),
        end_date=_as_date(cell_value("end_date"), defaults.end_date),
        ayanamsa=text("ayanamsa", defaults.ayanamsa),
        default_timezone=text("default_timezone", defaults.default_timezone),
        default_dst=text("default_dst", defaults.default_dst),
        event_location_name=text("event_location_name", defaults.event_location_name),
        event_latitude=number("event_latitude"),
        event_longitude=number("event_longitude"),
        event_timezone=text("event_timezone", ""),
        event_dst=text("event_dst", defaults.event_dst),
        user_name=text("user_name", defaults.user_name),
        natal_date=text("natal_date", defaults.natal_date),
        natal_time=text("natal_time", defaults.natal_time),
        natal_birth_place=text("natal_birth_place", defaults.natal_birth_place),
        natal_latitude=text("natal_latitude", defaults.natal_latitude),
        natal_longitude=text("natal_longitude", defaults.natal_longitude),
        natal_timezone=text("natal_timezone", defaults.natal_timezone),
        natal_moon_sign=text("natal_moon_sign", defaults.natal_moon_sign),
        natal_nakshatra=text("natal_nakshatra", defaults.natal_nakshatra),
        natal_lagna=text("natal_lagna", defaults.natal_lagna),
    )


def validate_foundation_config(config: FoundationConfig) -> None:
    if config.end_date < config.from_date:
        raise ValueError("End Date must be on or after From Date.")
    if config.event_latitude in ("", None) or config.event_longitude in ("", None) or not str(config.event_timezone).strip():
        raise ValueError(
            "Event location is required for muhurat calculation. Please set EventLatitude, EventLongitude, and EventTimezone."
        )


def build_event_location(config: FoundationConfig) -> LocationConfig:
    return LocationConfig(
        name=config.event_location_name or "Event Location",
        latitude=float(config.event_latitude),
        longitude=float(config.event_longitude),
        timezone=config.event_timezone,
        altitude_m=DEFAULT_LOCATION.altitude_m,
    )


def build_location_aware_raw_rows(config: FoundationConfig):
    """Preserve raw ephemeris logic while using event location for muhurat."""

    validate_foundation_config(config)

    event_location = build_event_location(config)
    astro = SwissEphemerisEngine(event_location)
    rows = []
    previous_signature = None
    final_next_sunrise = None

    for offset in range((config.end_date - config.from_date).days + 1):
        local_day = config.from_date + timedelta(days=offset)
        day_record = build_day_payload(local_day, astro)
        horas = day_record["horas"]
        choghadiya = day_record["choghadiya"]
        rahu_kaal = day_record["rahu_kaal"]
        yamaganda = day_record["yamaganda"]
        abhijit = day_record["abhijit_muhurat"]
        sunrise = floor_minute(pd.Timestamp(day_record["sunrise"]))
        sunset = floor_minute(pd.Timestamp(day_record["sunset"]))
        next_sunrise = floor_minute(pd.Timestamp(day_record["next_sunrise"]))
        final_next_sunrise = next_sunrise.to_pydatetime().replace(tzinfo=None)
        minute_ts = sunrise

        while minute_ts < next_sunrise:
            start_dt = minute_ts.to_pydatetime()
            end_guess = (minute_ts + timedelta(minutes=1)).to_pydatetime()
            snapshot = astro.planetary_snapshot(start_dt)
            panchang = astro.panchang_snapshot(start_dt)
            planets = snapshot["planets"]
            sun_lon = float(planets["Sun"]["longitude"])
            moon_lon = float(planets["Moon"]["longitude"])
            moon_parts = longitude_parts(moon_lon)
            lagna_parts = longitude_parts(float(panchang["ascendant_longitude"]))
            tithi = tithi_details(sun_lon, moon_lon)
            yoga = yoga_name(sun_lon, moon_lon)
            karana = karana_name(sun_lon, moon_lon)
            jd = local_datetime_to_jd(start_dt.replace(tzinfo=minute_ts.tzinfo))
            ayanamsa_value = round(swe.get_ayanamsa_ut(jd), 6)
            hora_segment = find_segment(start_dt, horas)
            choghadiya_segment = find_segment(start_dt, choghadiya)

            row = {
                "Date": start_dt.date(),
                "Day": str(day_record["weekday"]),
                "Start": start_dt.time(),
                "End": None,
                "Min": None,
                "StartDateTime": start_dt.replace(tzinfo=None),
                "EndDateTime": None,
                "Sunrise": sunrise.to_pydatetime().replace(tzinfo=None),
                "Sunset": sunset.to_pydatetime().replace(tzinfo=None),
                "Timezone": config.event_timezone,
                "DST": config.event_dst,
                "Ayanamsa": config.ayanamsa,
                "AyanamsaDeg": ayanamsa_value,
                "Paksha": tithi["paksha"],
                "Tithi": tithi["tithi"],
                "TithiNo": tithi["tithi_number"],
                "MoonNakshatra": moon_parts["nakshatra"],
                "MoonPada": moon_parts["pada"],
                "LagnaNakshatra": lagna_parts["nakshatra"],
                "LagnaPada": lagna_parts["pada"],
                "Yoga": yoga,
                "Karana": karana,
                "Choghadiya": str(choghadiya_segment["type"]),
                "Hora": str(hora_segment["ruler"]),
                "Abhijit": "Y" if overlaps(start_dt, end_guess, abhijit) else "N",
                "RahuKaal": "Y" if overlaps(start_dt, end_guess, rahu_kaal) else "N",
                "Yamaganda": "Y" if overlaps(start_dt, end_guess, yamaganda) else "N",
                "Gulika": "N",
                "Durmuhurta": "N",
                "Varjyam": "N",
                "TithiKshaya": "N",
                "NakshatraKshaya": "N",
                "LagnaSign": lagna_parts["sign"],
                "LagnaDeg": lagna_parts["degree"],
                "MoonSign": moon_parts["sign"],
                "MoonDeg": moon_parts["degree"],
                "MoonHouse": whole_sign_house(lagna_parts["sign"], moon_parts["sign"]),
                "MoonRetro": "Y" if planets["Moon"].get("retrograde") else "N",
            }

            for planet in ["Sun", "Mars", "Mercury", "Jupiter", "Venus", "Saturn", "Rahu", "Ketu"]:
                parts = longitude_parts(float(planets[planet]["longitude"]))
                row[f"{planet}Sign"] = parts["sign"]
                row[f"{planet}Deg"] = parts["degree"]
                row[f"{planet}House"] = whole_sign_house(lagna_parts["sign"], parts["sign"])
                row[f"{planet}Nakshatra"] = parts["nakshatra"]
                row[f"{planet}Pada"] = parts["pada"]
                row[f"{planet}Retro"] = "Y" if planets[planet].get("retrograde") else "N"

            signature_fields = [
                header
                for header in RAW_HEADERS
                if header
                not in {"Date", "Start", "End", "Min", "StartDateTime", "EndDateTime", "Sunrise", "Sunset", "AyanamsaDeg"}
                and not header.endswith("Deg")
            ]
            signature = tuple(row.get(field) for field in signature_fields)
            if signature == previous_signature:
                minute_ts += timedelta(minutes=1)
                continue

            previous_signature = signature
            rows.append(row)
            minute_ts += timedelta(minutes=1)

    for index, row in enumerate(rows):
        if index + 1 < len(rows):
            end_dt = rows[index + 1]["StartDateTime"]
        else:
            end_dt = final_next_sunrise or datetime.combine(config.end_date + timedelta(days=1), time.min)
        row["EndDateTime"] = end_dt
        row["End"] = end_dt.time()
        row["Min"] = max(0, round((end_dt - row["StartDateTime"]).total_seconds() / 60.0))

    return rows


def write_config_inputs(ws, config: FoundationConfig) -> None:
    clear_sheet(ws)
    ws.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    ws["A1"] = "Muhurat Finder V06 Workbook"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "V06 workbook base: raw ephemeris + event location + natal reference + helper columns. "
        "Parent-state columns are added in-place by the local V06 rule engine."
    )

    for cell in ["A3", "D3", "A10", "A22", "E22", "I22", "M22", "R22", "R53", "X53"]:
        ws[cell].fill = title_fill
        ws[cell].font = Font(color="FFFFFF", bold=True)

    ws["A3"] = "General Settings"
    general_items = [
        ("A4", "FromDate", CONFIG_CELL_MAP["from_date"], config.from_date),
        ("A5", "EndDate", CONFIG_CELL_MAP["end_date"], config.end_date),
        ("A6", "Ayanamsa", CONFIG_CELL_MAP["ayanamsa"], config.ayanamsa),
        ("A7", "DefaultTimezone", CONFIG_CELL_MAP["default_timezone"], config.default_timezone),
        ("A8", "DefaultDST", CONFIG_CELL_MAP["default_dst"], config.default_dst),
    ]
    for label_cell, label, value_cell, value in general_items:
        ws[label_cell] = label
        ws[value_cell] = value

    ws["D3"] = "Event / Current Muhurat Location"
    event_items = [
        ("D4", "EventLocationName", CONFIG_CELL_MAP["event_location_name"], config.event_location_name),
        ("D5", "EventLatitude", CONFIG_CELL_MAP["event_latitude"], config.event_latitude),
        ("D6", "EventLongitude", CONFIG_CELL_MAP["event_longitude"], config.event_longitude),
        ("D7", "EventTimezone", CONFIG_CELL_MAP["event_timezone"], config.event_timezone),
        ("D8", "EventDST", CONFIG_CELL_MAP["event_dst"], config.event_dst),
    ]
    for label_cell, label, value_cell, value in event_items:
        ws[label_cell] = label
        ws[value_cell] = value

    ws["A10"] = "Natal / Native Reference"
    natal_items = [
        ("A11", "UserName", CONFIG_CELL_MAP["user_name"], config.user_name),
        ("A12", "NatalDate", CONFIG_CELL_MAP["natal_date"], config.natal_date),
        ("A13", "NatalTime", CONFIG_CELL_MAP["natal_time"], config.natal_time),
        ("A14", "NatalBirthPlace", CONFIG_CELL_MAP["natal_birth_place"], config.natal_birth_place),
        ("A15", "NatalLatitude", CONFIG_CELL_MAP["natal_latitude"], config.natal_latitude),
        ("A16", "NatalLongitude", CONFIG_CELL_MAP["natal_longitude"], config.natal_longitude),
        ("A17", "NatalTimezone", CONFIG_CELL_MAP["natal_timezone"], config.natal_timezone),
        ("A18", "NatalMoonSign", CONFIG_CELL_MAP["natal_moon_sign"], config.natal_moon_sign),
        ("A19", "NatalNakshatra", CONFIG_CELL_MAP["natal_nakshatra"], config.natal_nakshatra),
        ("A20", "NatalLagna", CONFIG_CELL_MAP["natal_lagna"], config.natal_lagna),
    ]
    for label_cell, label, value_cell, value in natal_items:
        ws[label_cell] = label
        ws[value_cell] = value

    ws["D10"] = "Notes"
    ws["D10"].fill = section_fill
    ws["D10"].font = Font(bold=True)
    ws["D11"] = (
        "Birth location is used for natal reference. Event location is used for muhurat timing, "
        "Lagna, houses, sunrise/sunset, Hora, Choghadiya, Rahu Kaal, Yamaganda, Gulika, and Abhijit."
    )
    ws["D12"] = (
        "Changing EventLatitude/EventLongitude/EventTimezone after workbook generation does not recalculate raw astronomical "
        "values. Regenerate workbook for a new event location."
    )
    ws["D13"] = (
        "Birth location affects natal chart. Event location affects muhurat. "
        "Do not calculate muhurat using natal birth location unless the person is physically there during the event."
    )
    for cell in ["D11", "D12", "D13"]:
        ws[cell].fill = note_fill
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")

    for cell in [CONFIG_CELL_MAP["from_date"], CONFIG_CELL_MAP["end_date"]]:
        ws[cell].number_format = "dd-mmm-yyyy"


def write_config_lookup_tables(ws) -> None:
    global CONFIG_META

    table_fill = PatternFill("solid", fgColor="D9EAF7")
    table_header_fill = PatternFill("solid", fgColor="4F81BD")
    table_header_font = Font(color="FFFFFF", bold=True)

    def write_table(start_col: int, start_row: int, title: str, headers: list[str], rows: list[tuple[Any, ...]]):
        ws.cell(start_row, start_col, title)
        ws.cell(start_row, start_col).fill = PatternFill("solid", fgColor="1F4E78")
        ws.cell(start_row, start_col).font = Font(color="FFFFFF", bold=True)
        for offset, header in enumerate(headers):
            cell = ws.cell(start_row + 1, start_col + offset, header)
            cell.fill = table_header_fill
            cell.font = table_header_font
        for row_offset, row in enumerate(rows, start=2):
            for col_offset, value in enumerate(row):
                ws.cell(start_row + row_offset, start_col + col_offset, value)
                ws.cell(start_row + row_offset, start_col + col_offset).fill = table_fill
        return {
            "title_row": start_row,
            "header_row": start_row + 1,
            "data_row_start": start_row + 2,
            "data_row_end": start_row + 1 + len(rows),
            "col_start": start_col,
            "col_end": start_col + len(headers) - 1,
        }

    sign_rows = [
        ("Aries", 1),
        ("Taurus", 2),
        ("Gemini", 3),
        ("Cancer", 4),
        ("Leo", 5),
        ("Virgo", 6),
        ("Libra", 7),
        ("Scorpio", 8),
        ("Sagittarius", 9),
        ("Capricorn", 10),
        ("Aquarius", 11),
        ("Pisces", 12),
    ]
    sign_lord_rows = [
        ("Aries", 1, "Mars"),
        ("Taurus", 2, "Venus"),
        ("Gemini", 3, "Mercury"),
        ("Cancer", 4, "Moon"),
        ("Leo", 5, "Sun"),
        ("Virgo", 6, "Mercury"),
        ("Libra", 7, "Venus"),
        ("Scorpio", 8, "Mars"),
        ("Sagittarius", 9, "Jupiter"),
        ("Capricorn", 10, "Saturn"),
        ("Aquarius", 11, "Saturn"),
        ("Pisces", 12, "Jupiter"),
    ]
    nakshatra_rows = [
        ("Ashwini", 1, "Ketu"),
        ("Bharani", 2, "Venus"),
        ("Krittika", 3, "Sun"),
        ("Rohini", 4, "Moon"),
        ("Mrigashira", 5, "Mars"),
        ("Ardra", 6, "Rahu"),
        ("Punarvasu", 7, "Jupiter"),
        ("Pushya", 8, "Saturn"),
        ("Ashlesha", 9, "Mercury"),
        ("Magha", 10, "Ketu"),
        ("Purva Phalguni", 11, "Venus"),
        ("Uttara Phalguni", 12, "Sun"),
        ("Hasta", 13, "Moon"),
        ("Chitra", 14, "Mars"),
        ("Swati", 15, "Rahu"),
        ("Vishakha", 16, "Jupiter"),
        ("Anuradha", 17, "Saturn"),
        ("Jyeshtha", 18, "Mercury"),
        ("Moola", 19, "Ketu"),
        ("Purva Ashadha", 20, "Venus"),
        ("Uttara Ashadha", 21, "Sun"),
        ("Shravana", 22, "Moon"),
        ("Dhanishta", 23, "Mars"),
        ("Shatabhisha", 24, "Rahu"),
        ("Purva Bhadrapada", 25, "Jupiter"),
        ("Uttara Bhadrapada", 26, "Saturn"),
        ("Revati", 27, "Mercury"),
    ]
    tara_rows = [
        (1, "Janma", "Mixed"),
        (2, "Sampat", "Good"),
        (3, "Vipat", "Difficult"),
        (4, "Kshema", "Good"),
        (5, "Pratyak", "Difficult"),
        (6, "Sadhana", "Good"),
        (7, "Naidhana", "Difficult"),
        (8, "Mitra", "Good"),
        (9, "Parama Mitra", "Excellent"),
    ]
    dignity_rows = [
        ("Sun", "Leo", "Aries", "Libra", "Leo"),
        ("Moon", "Cancer", "Taurus", "Scorpio", "Taurus"),
        ("Mars", "Aries,Scorpio", "Capricorn", "Cancer", "Aries"),
        ("Mercury", "Gemini,Virgo", "Virgo", "Pisces", "Virgo"),
        ("Jupiter", "Sagittarius,Pisces", "Cancer", "Capricorn", "Sagittarius"),
        ("Venus", "Taurus,Libra", "Pisces", "Virgo", "Libra"),
        ("Saturn", "Capricorn,Aquarius", "Libra", "Aries", "Aquarius"),
    ]
    nature_rows = [
        ("Sun", "Malefic"),
        ("Moon", "Benefic"),
        ("Mars", "Malefic"),
        ("Mercury", "Neutral"),
        ("Jupiter", "Benefic"),
        ("Venus", "Benefic"),
        ("Saturn", "Malefic"),
        ("Rahu", "Malefic"),
        ("Ketu", "Malefic"),
    ]
    house_rows = [
        (1, "Self, body, identity, initiative", "presence, confidence, beginning"),
        (2, "wealth, speech, family, stored value", "money, family talk, purchase"),
        (3, "effort, communication, courage, short travel", "messages, research, execution, movement"),
        (4, "home, comfort, property, inner stability", "home, property, vehicles, domestic matters"),
        (5, "intelligence, creativity, romance, speculation", "learning, creativity, relationship, investing"),
        (6, "service, obstacles, enemies, disease, problem solving", "debugging, discipline, conflict handling"),
        (7, "partnership, public dealing, marriage, agreements", "relationship, contract, public dealing"),
        (8, "sudden events, risk, secrecy, instability", "avoid major starts"),
        (9, "dharma, fortune, higher learning, blessings", "wisdom, prayer, higher guidance"),
        (10, "career, authority, karma, status", "boss, management, leadership"),
        (11, "gains, network, fulfillment, income", "wealth, gains, approval, success"),
        (12, "loss, expenses, isolation, foreign, sleep", "avoid unless spiritual/rest context"),
    ]

    tables = {
        "sign_order": write_table(1, 22, "SignOrderTable", ["Sign", "SignNo"], sign_rows),
        "sign_lord": write_table(5, 22, "SignLordTable", ["Sign", "SignNo", "Lord"], sign_lord_rows),
        "nakshatra": write_table(9, 22, "NakshatraTable", ["Nakshatra", "NakshatraNo", "Lord"], nakshatra_rows),
        "tara": write_table(13, 22, "TaraNameTable", ["TaraCycleNo", "TaraName", "TaraQuality"], tara_rows),
        "dignity": write_table(18, 22, "PlanetDignityTable", ["Planet", "OwnSigns", "ExaltedSign", "DebilitatedSign", "MooltrikonaSign"], dignity_rows),
        "nature": write_table(18, 53, "PlanetNatureTable", ["Planet", "Nature"], nature_rows),
        "house": write_table(24, 53, "HouseMeaningTable", ["House", "CoreMeaning", "UsefulFor"], house_rows),
    }

    def data_range(meta):
        return f"'CONFIG'!${excel_col(meta['col_start'])}${meta['data_row_start']}:${excel_col(meta['col_end'])}${meta['data_row_end']}"

    def abs_ref(cell_ref: str) -> str:
        col = "".join(ch for ch in cell_ref if ch.isalpha())
        row = "".join(ch for ch in cell_ref if ch.isdigit())
        return f"'CONFIG'!${col}${row}"

    CONFIG_META = {
        "EventLocationName": abs_ref(CONFIG_CELL_MAP["event_location_name"]),
        "EventLatitude": abs_ref(CONFIG_CELL_MAP["event_latitude"]),
        "EventLongitude": abs_ref(CONFIG_CELL_MAP["event_longitude"]),
        "EventTimezone": abs_ref(CONFIG_CELL_MAP["event_timezone"]),
        "EventDST": abs_ref(CONFIG_CELL_MAP["event_dst"]),
        "NatalBirthPlace": abs_ref(CONFIG_CELL_MAP["natal_birth_place"]),
        "NatalLatitude": abs_ref(CONFIG_CELL_MAP["natal_latitude"]),
        "NatalLongitude": abs_ref(CONFIG_CELL_MAP["natal_longitude"]),
        "NatalTimezone": abs_ref(CONFIG_CELL_MAP["natal_timezone"]),
        "NatalMoonSign": abs_ref(CONFIG_CELL_MAP["natal_moon_sign"]),
        "NatalNakshatra": abs_ref(CONFIG_CELL_MAP["natal_nakshatra"]),
        "NatalLagna": abs_ref(CONFIG_CELL_MAP["natal_lagna"]),
        "SignOrderRange": data_range(tables["sign_order"]),
        "SignOrderNameRange": f"'CONFIG'!$A${tables['sign_order']['data_row_start']}:$A${tables['sign_order']['data_row_end']}",
        "SignOrderNoRange": f"'CONFIG'!$B${tables['sign_order']['data_row_start']}:$B${tables['sign_order']['data_row_end']}",
        "SignLordRange": data_range(tables["sign_lord"]),
        "NakshatraRange": data_range(tables["nakshatra"]),
        "NakshatraNameRange": f"'CONFIG'!$I${tables['nakshatra']['data_row_start']}:$I${tables['nakshatra']['data_row_end']}",
        "NakshatraNoRange": f"'CONFIG'!$J${tables['nakshatra']['data_row_start']}:$J${tables['nakshatra']['data_row_end']}",
        "TaraRange": data_range(tables["tara"]),
        "PlanetDignityRange": data_range(tables["dignity"]),
    }


def remove_old_formula_columns(ws) -> None:
    header_map = get_header_map(ws)
    delete_cols = sorted((header_map[name] for name in OLD_FORMULA_COLUMNS if name in header_map), reverse=True)
    for col_idx in delete_cols:
        ws.delete_cols(col_idx)


def append_or_update_column(ws, header_name: str) -> int:
    header_map = get_header_map(ws)
    if header_name in header_map:
        return header_map[header_name]
    next_col = ws.max_column + 1
    ws.cell(1, next_col, header_name)
    return next_col


def foundation_formula_map(row_number: int, header_map: dict[str, int], formula_col_map: dict[str, int]) -> dict[str, str]:
    def ref(header_name: str) -> str:
        col = formula_col_map.get(header_name, header_map.get(header_name))
        if col is None:
            raise KeyError(f"Missing column for formula: {header_name}")
        return f"{excel_col(col)}{row_number}"

    def sign_lookup(sign_ref: str) -> str:
        return f'IFERROR(VLOOKUP({sign_ref},{CONFIG_META["SignOrderRange"]},2,FALSE),"")'

    def house_sign_formula(house_no: int) -> str:
        sign_no = f"MOD({sign_lookup(ref('LagnaSign'))}+{house_no}-2,12)+1"
        return f'=IFERROR(INDEX({CONFIG_META["SignOrderNameRange"]},MATCH({sign_no},{CONFIG_META["SignOrderNoRange"]},0)),"")'

    def dignity_formula(planet: str) -> str:
        sign_ref = ref(f"{planet}Sign")
        own = f'VLOOKUP("{planet}",{CONFIG_META["PlanetDignityRange"]},2,FALSE)'
        exalted = f'VLOOKUP("{planet}",{CONFIG_META["PlanetDignityRange"]},3,FALSE)'
        debilitated = f'VLOOKUP("{planet}",{CONFIG_META["PlanetDignityRange"]},4,FALSE)'
        moola = f'VLOOKUP("{planet}",{CONFIG_META["PlanetDignityRange"]},5,FALSE)'
        return (
            f'=IF({sign_ref}="","",'
            f'IF({sign_ref}={exalted},"Exalted",'
            f'IF({sign_ref}={debilitated},"Debilitated",'
            f'IF({sign_ref}={moola},"Mooltrikona",'
            f'IF(ISNUMBER(SEARCH({sign_ref}&",",{own}&",")),"Own","Neutral")))))'
        )

    formulas = {
        "EventLocationName": f"={CONFIG_META['EventLocationName']}",
        "EventLatitude": f"={CONFIG_META['EventLatitude']}",
        "EventLongitude": f"={CONFIG_META['EventLongitude']}",
        "EventTimezone": f"={CONFIG_META['EventTimezone']}",
        "EventDST": f"={CONFIG_META['EventDST']}",
        "NatalBirthPlace": f"={CONFIG_META['NatalBirthPlace']}",
        "NatalLatitude": f"={CONFIG_META['NatalLatitude']}",
        "NatalLongitude": f"={CONFIG_META['NatalLongitude']}",
        "NatalTimezone": f"={CONFIG_META['NatalTimezone']}",
        "NatalMoonSign": f"={CONFIG_META['NatalMoonSign']}",
        "NatalNakshatra": f"={CONFIG_META['NatalNakshatra']}",
        "NatalLagna": f"={CONFIG_META['NatalLagna']}",
        "CurrentMoonNakshatraNo": f'=IFERROR(VLOOKUP({ref("MoonNakshatra")},{CONFIG_META["NakshatraRange"]},2,FALSE),"")',
        "ChandrabalamSignDistance": (
            f'=IF(OR({ref("NatalMoonSign")}="",{ref("MoonSign")}=""),"",'
            f'MOD(VLOOKUP({ref("MoonSign")},{CONFIG_META["SignOrderRange"]},2,FALSE)-'
            f'VLOOKUP({ref("NatalMoonSign")},{CONFIG_META["SignOrderRange"]},2,FALSE),12)+1)'
        ),
        "TarabalamNakshatraDistance": (
            f'=IF(OR({ref("NatalNakshatra")}="",{ref("CurrentMoonNakshatraNo")}=""),"",'
            f'MOD({ref("CurrentMoonNakshatraNo")}-VLOOKUP({ref("NatalNakshatra")},{CONFIG_META["NakshatraRange"]},2,FALSE),27)+1)'
        ),
        "TaraCycleNo": f'=IF({ref("TarabalamNakshatraDistance")}="","",MOD({ref("TarabalamNakshatraDistance")}-1,9)+1)',
        "TaraName": f'=IFERROR(VLOOKUP({ref("TaraCycleNo")},{CONFIG_META["TaraRange"]},2,FALSE),"")',
        "LagnaLord": f'=IFERROR(VLOOKUP({ref("LagnaSign")},{CONFIG_META["SignLordRange"]},3,FALSE),"")',
        "LocationValidityNote": (
            f'=IF(OR({ref("EventLocationName")}="",{ref("EventLatitude")}="",{ref("EventLongitude")}="",{ref("EventTimezone")}=""),'
            f'"Missing event location","Calculated for event location")'
        ),
        "CalculationLayerNote": '="V06 workbook base: raw ephemeris + location + natal + helper columns. Parent-state columns are applied later in the same workbook."',
    }

    for house_no in range(2, 13):
        sign_header = f"House{house_no}Sign"
        lord_header = f"House{house_no}Lord"
        formulas[sign_header] = house_sign_formula(house_no)
        formulas[lord_header] = f'=IFERROR(VLOOKUP({ref(sign_header)},{CONFIG_META["SignLordRange"]},3,FALSE),"")'

    for planet in ["Sun", "Moon", "Mars", "Mercury", "Jupiter", "Venus", "Saturn"]:
        formulas[f"{planet}Dignity"] = dignity_formula(planet)

    return formulas


def apply_foundation_formulas(ws, config_ws) -> None:
    del config_ws  # Header-based references are driven by CONFIG_META.

    remove_old_formula_columns(ws)
    header_map = get_header_map(ws)
    formula_col_map = {}
    for header in FOUNDATION_COLUMNS:
        formula_col_map[header] = append_or_update_column(ws, header)
    header_map = get_header_map(ws)

    for row_number in range(2, ws.max_row + 1):
        formulas = foundation_formula_map(row_number, header_map, formula_col_map)
        for header, formula in formulas.items():
            ws.cell(row_number, formula_col_map[header], formula)


def apply_formatting(wb: Workbook) -> None:
    config_ws = wb["CONFIG"]
    ephemeris_ws = wb["EPHEMERIS_RAW"]

    config_ws.freeze_panes = "A2"
    for col in range(1, 33):
        width = 16
        if col in (1, 4, 9, 13, 18, 24, 28):
            width = 20
        config_ws.column_dimensions[excel_col(col)].width = width

    ephemeris_ws.freeze_panes = "A2"
    ephemeris_ws.auto_filter.ref = ephemeris_ws.dimensions
    header_map = get_header_map(ephemeris_ws)
    for cell in ephemeris_ws[1]:
        cell.font = Font(bold=True)

    fills = {
        "event": PatternFill("solid", fgColor="D9EAF7"),
        "natal": PatternFill("solid", fgColor="EADCF8"),
        "tara": PatternFill("solid", fgColor="FFF2CC"),
        "house": PatternFill("solid", fgColor="E2F0D9"),
        "dignity": PatternFill("solid", fgColor="FCE4D6"),
        "note": PatternFill("solid", fgColor="E7E6E6"),
    }

    groups = [
        (EVENT_COLUMNS, "event"),
        (NATAL_COLUMNS, "natal"),
        (TARA_CHANDRA_COLUMNS, "tara"),
        (HOUSE_COLUMNS, "house"),
        (DIGNITY_COLUMNS, "dignity"),
        (NOTE_COLUMNS, "note"),
    ]

    for headers, fill_key in groups:
        for header in headers:
            if header in header_map:
                cell = ephemeris_ws.cell(1, header_map[header])
                cell.fill = fills[fill_key]

    for col_idx in range(1, ephemeris_ws.max_column + 1):
        header = ephemeris_ws.cell(1, col_idx).value
        width = min(max(len(str(header or "")) + 2, 10), 24)
        ephemeris_ws.column_dimensions[excel_col(col_idx)].width = width


def build_initial_config() -> FoundationConfig:
    defaults = default_foundation_config()
    if OUTPUT_WORKBOOK.exists():
        return read_foundation_config(OUTPUT_WORKBOOK)
    return defaults


def main() -> None:
    config = build_initial_config()
    rows = build_location_aware_raw_rows(config)

    wb = Workbook()
    config_ws = wb.active
    config_ws.title = "CONFIG"
    write_config_inputs(config_ws, config)
    write_config_lookup_tables(config_ws)

    raw_ws = get_or_create_sheet(wb, "EPHEMERIS_RAW")
    write_sheet(raw_ws, rows, RAW_HEADERS)
    apply_foundation_formulas(raw_ws, config_ws)

    try:
        from openpyxl.workbook.properties import CalcProperties

        wb.calculation = CalcProperties(calcMode="auto")
    except Exception:
        pass

    apply_formatting(wb)
    wb.save(OUTPUT_WORKBOOK)
    print(OUTPUT_WORKBOOK)


if __name__ == "__main__":
    main()
