"""Apply the V05-style parent-state rule engine in-place on the V06 workbook."""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Sequence
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
BUILD_DIR = SCRIPT_DIR.parent
BASE_WORKBOOK = BUILD_DIR / "MuhuratFinder_V06_Workbook.xlsx"
DEBUG_DIR = BUILD_DIR / "debug"
OUTPUT_WORKBOOK = BASE_WORKBOOK

OLD_SCORING_COLUMNS = [
    "TithiScore",
    "NakshatraScore",
    "YogaScore",
    "KaranaScore",
    "PakshaScore",
    "ChoghadiyaScore",
    "HoraScore",
    "AbhijitScore",
    "RahuPenalty",
    "YamagandaPenalty",
    "GulikaPenalty",
    "DurmuhurtaPenalty",
    "VarjyamPenalty",
    "TithiKshayaPenalty",
    "NakshatraKshayaPenalty",
    "LagnaDegreeScore",
    "LagnaSignScore",
    "LagnaNakshatraScore",
    "MaleficPressureScore",
    "MoonHouseScore",
    "MoonDegreeScore",
    "MoonNakshatraQualityScore",
    "VenusStrengthScore",
    "JupiterSupportScore",
    "MercuryCommunicationScore",
    "ChandrabalamScore",
    "TarabalamScore",
    "PanchangScore",
    "TimeFilterScore",
    "LagnaStrengthScore",
    "MoonStrengthScore",
    "PlanetarySupportScore",
    "PersonalCompatibilityScore",
    "OverallMuhuratScore",
]

V05_COLUMNS = [
    "IsRahuKaal",
    "IsYamaganda",
    "IsGulika",
    "IsDurmuhurta",
    "IsVarjyam",
    "IsTithiKshaya",
    "IsNakshatraKshaya",
    "IsHarshYoga",
    "IsWeakMoonRisk",
    "AvoidScore",
    "GoldenScore",
    "AuspiciousScore",
    "LeadershipScore",
    "WealthScore",
    "RelationshipScore",
    "LearningScore",
    "ExecutionScore",
    "TravelScore",
    "PurchaseScore",
    "PrimaryState",
    "PrimaryStateReason",
    "SecondaryState1",
    "SecondaryState2",
    "SecondaryState3",
    "SecondaryStates",
    "SecondaryStateReason",
    "RiskLevel",
    "RiskReason",
    "BestActions",
    "AvoidActions",
    "V05CalculationNote",
]

FLAG_COLUMNS = V05_COLUMNS[:9]
SCORE_COLUMNS = V05_COLUMNS[9:19]
OUTPUT_COLUMNS = V05_COLUMNS[19:]
FORBIDDEN_SHEETS = {"INDICATORS", "DASHBOARD", "DAY_SUMMARY", "SHORTLIST", "RECOMMENDATIONS"}
REQUIRED_OUTPUT_HEADERS = ["PrimaryState", "RiskLevel", "BestActions", "AvoidActions"]
MAX_SUSPICIOUS_PRINT = 25


def excel_col(col_index: int) -> str:
    return get_column_letter(col_index)


def get_header_map(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value not in (None, "")
    }


def safe_col(header_map: dict[str, int], header_name: str) -> int:
    if header_name not in header_map:
        raise KeyError(f"Header not found: {header_name}")
    return header_map[header_name]


def append_or_update_column(ws, header_name: str) -> int:
    header_map = get_header_map(ws)
    if header_name in header_map:
        return header_map[header_name]
    new_col = ws.max_column + 1
    ws.cell(1, new_col, header_name)
    return new_col


def clear_conditional_formatting(ws) -> None:
    """Remove all conditional formatting rules to isolate XML/formula issues."""
    if hasattr(ws.conditional_formatting, "_cf_rules"):
        ws.conditional_formatting._cf_rules.clear()


def remove_old_formula_columns(ws) -> None:
    header_map = get_header_map(ws)
    delete_cols = sorted((header_map[h] for h in OLD_SCORING_COLUMNS if h in header_map), reverse=True)
    for col_idx in delete_cols:
        ws.delete_cols(col_idx)


def remove_existing_v05_columns(ws) -> None:
    header_map = get_header_map(ws)
    delete_cols = sorted((header_map[h] for h in V05_COLUMNS if h in header_map), reverse=True)
    for col_idx in delete_cols:
        ws.delete_cols(col_idx)


def contains_formula(ref: str, text: str) -> str:
    return f'ISNUMBER(SEARCH("{text}",{ref}))'


def yes_formula(ref: str) -> str:
    return f'IF(OR({ref}="Y",{ref}="Yes"),1,0)'


def build_formulas(row_number: int, header_map: dict[str, int], v05_col_map: dict[str, int]) -> dict[str, str]:
    def ref(header_name: str) -> str:
        col = v05_col_map.get(header_name, header_map.get(header_name))
        if col is None:
            raise KeyError(f"Missing header in formula builder: {header_name}")
        return f"{excel_col(col)}{row_number}"

    choghadiya = ref("Choghadiya")
    hora = ref("Hora")
    yoga = ref("Yoga")
    moon_house = ref("MoonHouse")
    tara_name = ref("TaraName")
    chandrabalam = ref("ChandrabalamSignDistance")
    lagna_deg = ref("LagnaDeg")

    harsh_yoga_or = (
        f'OR({yoga}="Vyaghata",{yoga}="Atiganda",{yoga}="Shoola",{yoga}="Ganda",{yoga}="Vyatipata",{yoga}="Vaidhriti")'
    )
    weak_moon_or = (
        f'OR({moon_house}=8,{moon_house}=12,AND({moon_house}=6,OR({tara_name}="Vipat",{tara_name}="Pratyak",{tara_name}="Naidhana")))'
    )
    avoid_hard_or = (
        f'OR({ref("IsRahuKaal")}=1,{ref("IsYamaganda")}=1,{ref("IsGulika")}=1,{ref("IsDurmuhurta")}=1,'
        f'{ref("IsVarjyam")}=1,{ref("IsTithiKshaya")}=1,{ref("IsNakshatraKshaya")}=1)'
    )

    golden_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'20*IF({contains_formula(choghadiya,"Amrit")},1,0)+'
        f'15*IF({contains_formula(choghadiya,"Labh")},1,0)+'
        f'10*IF(OR({contains_formula(choghadiya,"Shubh")},{contains_formula(choghadiya,"Shubha")}),1,0)+'
        f'10*IF(OR({contains_formula(hora,"Jupiter")},{contains_formula(hora,"Venus")},{contains_formula(hora,"Mercury")},{contains_formula(hora,"Moon")}),1,0)+'
        f'10*IF(OR({moon_house}=1,{moon_house}=5,{moon_house}=9,{moon_house}=10,{moon_house}=11),1,0)+'
        f'10*IF(OR({tara_name}="Mitra",{tara_name}="Parama Mitra"),1,0)+'
        f'10*IF(OR({chandrabalam}=1,{chandrabalam}=3,{chandrabalam}=6,{chandrabalam}=7,{chandrabalam}=10,{chandrabalam}=11),1,0)+'
        f'10*IF(OR({ref("JupiterDignity")}="Exalted",{ref("JupiterDignity")}="Own",{ref("JupiterDignity")}="Mooltrikona"),1,0)+'
        f'10*IF(OR({ref("VenusDignity")}="Exalted",{ref("VenusDignity")}="Own",{ref("VenusDignity")}="Mooltrikona"),1,0)+'
        f'10*IF(AND({lagna_deg}>=5,{lagna_deg}<=25),1,0)-'
        f'20*IF({ref("IsHarshYoga")}=1,1,0)-'
        f'15*IF({ref("IsWeakMoonRisk")}=1,1,0))'
    )
    auspicious_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'15*IF(OR({contains_formula(choghadiya,"Amrit")},{contains_formula(choghadiya,"Labh")},{contains_formula(choghadiya,"Shubh")},{contains_formula(choghadiya,"Shubha")}),1,0)+'
        f'10*IF(OR({contains_formula(hora,"Jupiter")},{contains_formula(hora,"Venus")},{contains_formula(hora,"Mercury")},{contains_formula(hora,"Moon")}),1,0)+'
        f'10*IF(OR({moon_house}=1,{moon_house}=3,{moon_house}=5,{moon_house}=7,{moon_house}=9,{moon_house}=10,{moon_house}=11),1,0)+'
        f'10*IF(OR({tara_name}="Sampat",{tara_name}="Kshema",{tara_name}="Sadhana",{tara_name}="Mitra",{tara_name}="Parama Mitra"),1,0)+'
        f'10*IF(OR({chandrabalam}=1,{chandrabalam}=3,{chandrabalam}=6,{chandrabalam}=7,{chandrabalam}=10,{chandrabalam}=11),1,0)+'
        f'5*IF(AND({lagna_deg}>=5,{lagna_deg}<=25),1,0)+'
        f'5*IF({ref("IsHarshYoga")}=0,1,0)-'
        f'15*IF({ref("IsHarshYoga")}=1,1,0)-'
        f'10*IF({ref("IsWeakMoonRisk")}=1,1,0))'
    )
    leadership_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'20*IF(OR({ref("SunHouse")}=1,{ref("SunHouse")}=10,{ref("SunHouse")}=11),1,0)+'
        f'15*IF(OR({ref("JupiterHouse")}=9,{ref("JupiterHouse")}=10,{ref("JupiterHouse")}=11),1,0)+'
        f'15*IF(OR({ref("MercuryHouse")}=3,{ref("MercuryHouse")}=10,{ref("MercuryHouse")}=11),1,0)+'
        f'10*IF(OR({ref("House10Lord")}="Sun",{ref("House10Lord")}="Mars",{ref("House10Lord")}="Jupiter",{ref("House10Lord")}="Mercury"),1,0)+'
        f'10*IF(OR({contains_formula(hora,"Sun")},{contains_formula(hora,"Jupiter")},{contains_formula(hora,"Mercury")},{contains_formula(hora,"Mars")}),1,0)+'
        f'10*IF(OR({contains_formula(choghadiya,"Labh")},{contains_formula(choghadiya,"Shubh")},{contains_formula(choghadiya,"Shubha")}),1,0)+'
        f'10*IF(OR({ref("SunDignity")}="Exalted",{ref("SunDignity")}="Own",{ref("SunDignity")}="Mooltrikona"),1,0)+'
        f'5*IF(AND({lagna_deg}>=5,{lagna_deg}<=25),1,0)-'
        f'15*IF({ref("IsHarshYoga")}=1,1,0)-'
        f'10*IF(OR({moon_house}=8,{moon_house}=12),1,0))'
    )
    wealth_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'20*IF(OR({ref("VenusHouse")}=2,{ref("VenusHouse")}=11),1,0)+'
        f'20*IF(OR({ref("JupiterHouse")}=2,{ref("JupiterHouse")}=11),1,0)+'
        f'15*IF(OR({ref("MercuryHouse")}=2,{ref("MercuryHouse")}=10,{ref("MercuryHouse")}=11),1,0)+'
        f'15*IF({ref("RahuHouse")}=11,1,0)+'
        f'15*IF({contains_formula(choghadiya,"Labh")},1,0)+'
        f'10*IF(OR({contains_formula(hora,"Venus")},{contains_formula(hora,"Jupiter")},{contains_formula(hora,"Mercury")}),1,0)+'
        f'10*IF(OR({ref("House2Lord")}="Venus",{ref("House2Lord")}="Jupiter",{ref("House2Lord")}="Mercury"),1,0)+'
        f'10*IF(OR({ref("House11Lord")}="Venus",{ref("House11Lord")}="Jupiter",{ref("House11Lord")}="Mercury",{ref("House11Lord")}="Saturn"),1,0)+'
        f'10*IF(OR({ref("JupiterDignity")}="Exalted",{ref("JupiterDignity")}="Own",{ref("JupiterDignity")}="Mooltrikona"),1,0)+'
        f'10*IF(OR({ref("VenusDignity")}="Exalted",{ref("VenusDignity")}="Own",{ref("VenusDignity")}="Mooltrikona"),1,0)-'
        f'15*IF({ref("IsHarshYoga")}=1,1,0)-'
        f'15*IF(OR({moon_house}=8,{moon_house}=12),1,0))'
    )
    relationship_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'20*IF(OR({ref("VenusHouse")}=1,{ref("VenusHouse")}=2,{ref("VenusHouse")}=5,{ref("VenusHouse")}=7,{ref("VenusHouse")}=11),1,0)+'
        f'15*IF(OR({ref("JupiterHouse")}=2,{ref("JupiterHouse")}=5,{ref("JupiterHouse")}=7,{ref("JupiterHouse")}=9,{ref("JupiterHouse")}=11),1,0)+'
        f'15*IF(OR({moon_house}=1,{moon_house}=2,{moon_house}=5,{moon_house}=7,{moon_house}=9,{moon_house}=11),1,0)+'
        f'10*IF(OR({ref("House7Lord")}="Venus",{ref("House7Lord")}="Jupiter",{ref("House7Lord")}="Moon",{ref("House7Lord")}="Mercury"),1,0)+'
        f'10*IF(OR({contains_formula(hora,"Venus")},{contains_formula(hora,"Jupiter")},{contains_formula(hora,"Moon")}),1,0)+'
        f'10*IF(OR({contains_formula(choghadiya,"Shubh")},{contains_formula(choghadiya,"Shubha")},{contains_formula(choghadiya,"Amrit")}),1,0)+'
        f'10*IF(OR({tara_name}="Mitra",{tara_name}="Parama Mitra"),1,0)+'
        f'10*IF(OR({chandrabalam}=1,{chandrabalam}=3,{chandrabalam}=6,{chandrabalam}=7,{chandrabalam}=10,{chandrabalam}=11),1,0)+'
        f'10*IF(OR({ref("VenusDignity")}="Exalted",{ref("VenusDignity")}="Own",{ref("VenusDignity")}="Mooltrikona"),1,0)-'
        f'20*IF({ref("IsHarshYoga")}=1,1,0)-'
        f'20*IF(OR({moon_house}=6,{moon_house}=8,{moon_house}=12),1,0)-'
        f'10*IF(OR({ref("KetuHouse")}=5,{ref("KetuHouse")}=7),1,0)-'
        f'10*IF({ref("MarsHouse")}=7,1,0))'
    )
    learning_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'20*IF(OR({ref("MercuryHouse")}=1,{ref("MercuryHouse")}=3,{ref("MercuryHouse")}=5,{ref("MercuryHouse")}=9,{ref("MercuryHouse")}=10,{ref("MercuryHouse")}=11),1,0)+'
        f'20*IF(OR({ref("JupiterHouse")}=1,{ref("JupiterHouse")}=4,{ref("JupiterHouse")}=5,{ref("JupiterHouse")}=9,{ref("JupiterHouse")}=10,{ref("JupiterHouse")}=11),1,0)+'
        f'10*IF(OR({ref("House5Lord")}="Mercury",{ref("House5Lord")}="Jupiter"),1,0)+'
        f'10*IF(OR({ref("House9Lord")}="Mercury",{ref("House9Lord")}="Jupiter"),1,0)+'
        f'10*IF(OR({contains_formula(hora,"Mercury")},{contains_formula(hora,"Jupiter")}),1,0)+'
        f'10*IF(OR({contains_formula(choghadiya,"Labh")},{contains_formula(choghadiya,"Shubh")},{contains_formula(choghadiya,"Shubha")},{contains_formula(choghadiya,"Amrit")}),1,0)+'
        f'10*IF(OR({ref("MoonNakshatra")}="Pushya",{ref("MoonNakshatra")}="Revati",{ref("MoonNakshatra")}="Hasta",{ref("MoonNakshatra")}="Rohini",{ref("MoonNakshatra")}="Ashwini",{ref("MoonNakshatra")}="Punarvasu",{ref("MoonNakshatra")}="Uttara Phalguni"),1,0)+'
        f'10*IF(OR({tara_name}="Sadhana",{tara_name}="Mitra",{tara_name}="Parama Mitra"),1,0)-'
        f'10*IF({ref("IsHarshYoga")}=1,1,0)-'
        f'10*IF(OR({moon_house}=8,{moon_house}=12),1,0))'
    )
    execution_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'20*IF(OR({ref("MercuryHouse")}=3,{ref("MercuryHouse")}=6,{ref("MercuryHouse")}=10,{ref("MercuryHouse")}=11),1,0)+'
        f'15*IF(OR({ref("MarsHouse")}=3,{ref("MarsHouse")}=6,{ref("MarsHouse")}=10,{ref("MarsHouse")}=11),1,0)+'
        f'15*IF(OR({ref("SaturnHouse")}=3,{ref("SaturnHouse")}=6,{ref("SaturnHouse")}=10,{ref("SaturnHouse")}=11),1,0)+'
        f'15*IF({moon_house}=6,1,0)+'
        f'10*IF(OR({ref("House6Lord")}="Mars",{ref("House6Lord")}="Mercury",{ref("House6Lord")}="Saturn"),1,0)+'
        f'10*IF(OR({ref("House10Lord")}="Mars",{ref("House10Lord")}="Mercury",{ref("House10Lord")}="Saturn"),1,0)+'
        f'10*IF(OR({contains_formula(hora,"Mercury")},{contains_formula(hora,"Mars")},{contains_formula(hora,"Saturn")}),1,0)+'
        f'10*IF(OR({contains_formula(choghadiya,"Chara")},{contains_formula(choghadiya,"Labh")},{contains_formula(choghadiya,"Shubh")},{contains_formula(choghadiya,"Shubha")}),1,0)-'
        f'10*IF({ref("IsHarshYoga")}=1,1,0)-'
        f'10*IF(OR({moon_house}=8,{moon_house}=12),1,0))'
    )
    travel_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'20*IF({contains_formula(choghadiya,"Chara")},1,0)+'
        f'15*IF(OR({ref("MarsHouse")}=3,{ref("MarsHouse")}=9),1,0)+'
        f'15*IF(OR({ref("MercuryHouse")}=3,{ref("MercuryHouse")}=9),1,0)+'
        f'10*IF(OR({moon_house}=3,{moon_house}=7,{moon_house}=9,{moon_house}=11),1,0)+'
        f'10*IF(OR({ref("House3Lord")}="Mars",{ref("House3Lord")}="Mercury",{ref("House3Lord")}="Moon",{ref("House3Lord")}="Jupiter"),1,0)+'
        f'10*IF(OR({ref("House9Lord")}="Mars",{ref("House9Lord")}="Mercury",{ref("House9Lord")}="Moon",{ref("House9Lord")}="Jupiter"),1,0)+'
        f'10*IF(OR({contains_formula(hora,"Mercury")},{contains_formula(hora,"Moon")},{contains_formula(hora,"Mars")}),1,0)+'
        f'10*IF(OR({ref("MoonNakshatra")}="Ashwini",{ref("MoonNakshatra")}="Pushya",{ref("MoonNakshatra")}="Hasta",{ref("MoonNakshatra")}="Anuradha",{ref("MoonNakshatra")}="Revati",{ref("MoonNakshatra")}="Uttara Ashadha"),1,0)-'
        f'20*IF(OR({moon_house}=8,{moon_house}=12),1,0)-'
        f'15*IF({ref("IsHarshYoga")}=1,1,0))'
    )
    purchase_raw = (
        f'IF({ref("AvoidScore")}>0,0,'
        f'20*IF(OR({ref("VenusHouse")}=2,{ref("VenusHouse")}=4,{ref("VenusHouse")}=11),1,0)+'
        f'15*IF(OR({ref("MercuryHouse")}=2,{ref("MercuryHouse")}=3,{ref("MercuryHouse")}=10,{ref("MercuryHouse")}=11),1,0)+'
        f'15*IF(OR({ref("JupiterHouse")}=2,{ref("JupiterHouse")}=9,{ref("JupiterHouse")}=11),1,0)+'
        f'15*IF(OR({contains_formula(choghadiya,"Labh")},{contains_formula(choghadiya,"Shubh")},{contains_formula(choghadiya,"Shubha")},{contains_formula(choghadiya,"Amrit")}),1,0)+'
        f'10*IF(OR({contains_formula(hora,"Venus")},{contains_formula(hora,"Mercury")},{contains_formula(hora,"Jupiter")}),1,0)+'
        f'10*IF(OR({ref("House2Lord")}="Venus",{ref("House2Lord")}="Mercury",{ref("House2Lord")}="Jupiter"),1,0)+'
        f'10*IF(OR({ref("House4Lord")}="Venus",{ref("House4Lord")}="Mercury",{ref("House4Lord")}="Jupiter",{ref("House4Lord")}="Moon"),1,0)+'
        f'10*IF(OR({ref("House11Lord")}="Venus",{ref("House11Lord")}="Mercury",{ref("House11Lord")}="Jupiter",{ref("House11Lord")}="Saturn"),1,0)+'
        f'10*IF(OR({ref("VenusDignity")}="Exalted",{ref("VenusDignity")}="Own",{ref("VenusDignity")}="Mooltrikona"),1,0)+'
        f'10*IF(OR({ref("MercuryDignity")}="Exalted",{ref("MercuryDignity")}="Own",{ref("MercuryDignity")}="Mooltrikona"),1,0)-'
        f'15*IF(OR({moon_house}=8,{moon_house}=12),1,0)-'
        f'15*IF({ref("IsHarshYoga")}=1,1,0))'
    )

    specific_refs = [
        ref("LeadershipScore"),
        ref("WealthScore"),
        ref("RelationshipScore"),
        ref("LearningScore"),
        ref("ExecutionScore"),
        ref("TravelScore"),
        ref("PurchaseScore"),
    ]
    max_specific = f"MAX({','.join(specific_refs)})"
    highest_specific_state = (
        f'IF({ref("LeadershipScore")}={max_specific},"Leadership / Authority",'
        f'IF({ref("WealthScore")}={max_specific},"Wealth / Gain",'
        f'IF({ref("RelationshipScore")}={max_specific},"Relationship / Social",'
        f'IF({ref("LearningScore")}={max_specific},"Learning / Wisdom",'
        f'IF({ref("ExecutionScore")}={max_specific},"Execution / Problem Solving",'
        f'IF({ref("TravelScore")}={max_specific},"Travel / Movement",'
        f'IF({ref("PurchaseScore")}={max_specific},"Purchase / Acquisition","")))))))'
    )

    candidates: list[tuple[str, str]] = [
        ("Golden Muhurat", f"{ref('GoldenScore')}>=75"),
        ("Auspicious Muhurat", f"{ref('AuspiciousScore')}>=50"),
        ("Leadership / Authority", f"{ref('LeadershipScore')}>=50"),
        ("Wealth / Gain", f"{ref('WealthScore')}>=50"),
        ("Relationship / Social", f"{ref('RelationshipScore')}>=50"),
        ("Learning / Wisdom", f"{ref('LearningScore')}>=50"),
        ("Execution / Problem Solving", f"{ref('ExecutionScore')}>=50"),
        ("Travel / Movement", f"{ref('TravelScore')}>=50"),
        ("Purchase / Acquisition", f"{ref('PurchaseScore')}>=50"),
    ]
    def nested_choice(exclude_refs: Sequence[str]) -> str:
        expr = '""'
        for name, cond in reversed(candidates):
            exclusions = "".join(f',{exclude}<>"{name}"' for exclude in exclude_refs)
            expr = f'IF(AND({cond},{ref("PrimaryState")}<>"{name}"{exclusions}),"{name}",{expr})'
        return expr

    secondary_1 = nested_choice(())
    secondary_2 = nested_choice((ref("SecondaryState1"),))
    secondary_3 = nested_choice((ref("SecondaryState1"), ref("SecondaryState2")))

    return {
        "IsRahuKaal": f"={yes_formula(ref('RahuKaal'))}",
        "IsYamaganda": f"={yes_formula(ref('Yamaganda'))}",
        "IsGulika": f"={yes_formula(ref('Gulika'))}",
        "IsDurmuhurta": f"={yes_formula(ref('Durmuhurta'))}",
        "IsVarjyam": f"={yes_formula(ref('Varjyam'))}",
        "IsTithiKshaya": f"={yes_formula(ref('TithiKshaya'))}",
        "IsNakshatraKshaya": f"={yes_formula(ref('NakshatraKshaya'))}",
        "IsHarshYoga": f'=IF({harsh_yoga_or},1,0)',
        "IsWeakMoonRisk": f'=IF({weak_moon_or},1,0)',
        "AvoidScore": f'=IF({avoid_hard_or},100,IF(AND({ref("IsHarshYoga")}=1,{ref("IsWeakMoonRisk")}=1),80,0))',
        "GoldenScore": f'=MAX(0,MIN(100,{golden_raw}))',
        "AuspiciousScore": f'=MAX(0,MIN(100,{auspicious_raw}))',
        "LeadershipScore": f'=MAX(0,MIN(100,{leadership_raw}))',
        "WealthScore": f'=MAX(0,MIN(100,{wealth_raw}))',
        "RelationshipScore": f'=MAX(0,MIN(100,{relationship_raw}))',
        "LearningScore": f'=MAX(0,MIN(100,{learning_raw}))',
        "ExecutionScore": f'=MAX(0,MIN(100,{execution_raw}))',
        "TravelScore": f'=MAX(0,MIN(100,{travel_raw}))',
        "PurchaseScore": f'=MAX(0,MIN(100,{purchase_raw}))',
        "PrimaryState": (
            f'=IF({ref("AvoidScore")}>=80,"Avoid / Bad Kaal",'
            f'IF({ref("GoldenScore")}>=75,"Golden Muhurat",'
            f'IF(AND({ref("AuspiciousScore")}>=60,{max_specific}<65),"Auspicious Muhurat",'
            f'IF({max_specific}>=50,{highest_specific_state},'
            f'IF({ref("AuspiciousScore")}>=50,"Auspicious Muhurat","Neutral / Routine")))))'
        ),
        "PrimaryStateReason": (
            f'=IF({ref("PrimaryState")}="Avoid / Bad Kaal","Blocked by active risk filters; keep work routine and reversible.",'
            f'IF({ref("PrimaryState")}="Golden Muhurat","Multiple strong supports align with low risk, making this a high-quality window.",'
            f'IF({ref("PrimaryState")}="Auspicious Muhurat","General support is positive and no major obstruction dominates the window.",'
            f'IF({ref("PrimaryState")}="Leadership / Authority","Authority, communication, visibility, or career-facing indicators are strongest now.",'
            f'IF({ref("PrimaryState")}="Wealth / Gain","Wealth, gain, value, or money-related indicators are strongest now.",'
            f'IF({ref("PrimaryState")}="Relationship / Social","Relationship, harmony, partnership, or social indicators are strongest now.",'
            f'IF({ref("PrimaryState")}="Learning / Wisdom","Study, research, judgment, or wisdom-oriented indicators are strongest now.",'
            f'IF({ref("PrimaryState")}="Execution / Problem Solving","Execution, service, logic, or problem-solving indicators are strongest now.",'
            f'IF({ref("PrimaryState")}="Travel / Movement","Movement, travel, visits, or transition-oriented indicators are strongest now.",'
            f'IF({ref("PrimaryState")}="Purchase / Acquisition","Acquisition, tools, assets, or practical purchase indicators are strongest now.","No single strong state dominates this window."))))))))))'
        ),
        "SecondaryState1": f"={secondary_1}",
        "SecondaryState2": f"={secondary_2}",
        "SecondaryState3": f"={secondary_3}",
        "SecondaryStates": (
            f'=IF({ref("SecondaryState1")}="","",'
            f'{ref("SecondaryState1")}&IF({ref("SecondaryState2")}="","","; "&{ref("SecondaryState2")})&IF({ref("SecondaryState3")}="","","; "&{ref("SecondaryState3")}))'
        ),
        "SecondaryStateReason": f'=IF({ref("SecondaryStates")}="","No strong secondary state.","Secondary states indicate additional supported action areas, but PrimaryState should be preferred.")',
        "RiskLevel": f'=IF({ref("AvoidScore")}>=80,"High Risk",IF(OR({ref("IsHarshYoga")}=1,{ref("IsWeakMoonRisk")}=1),"Medium Risk","Low Risk"))',
        "RiskReason": (
            f'=IF({ref("IsRahuKaal")}=1,"Rahu Kaal is active; avoid new starts.",'
            f'IF({ref("IsYamaganda")}=1,"Yamaganda is active; avoid initiating important actions.",'
            f'IF({ref("IsGulika")}=1,"Gulika is active; avoid major commitments.",'
            f'IF({ref("IsDurmuhurta")}=1,"Durmuhurta is active; avoid important starts.",'
            f'IF({ref("IsVarjyam")}=1,"Varjyam is active; avoid important actions.",'
            f'IF({ref("IsTithiKshaya")}=1,"Tithi Kshaya indicates unstable timing.",'
            f'IF({ref("IsNakshatraKshaya")}=1,"Nakshatra Kshaya indicates unstable timing.",'
            f'IF({ref("IsHarshYoga")}=1,"Harsh Yoga is active; use caution.",'
            f'IF({ref("IsWeakMoonRisk")}=1,"Moon placement is sensitive; avoid emotional or risky decisions.","No major risk filter active.")))))))))'
        ),
        "BestActions": (
            f'=IF({ref("PrimaryState")}="Golden Muhurat","Prayer, sankalp, important start, major positive commitment, long-term initiative.",'
            f'IF({ref("PrimaryState")}="Auspicious Muhurat","Start useful work, contact people, begin learning, make constructive decisions.",'
            f'IF({ref("PrimaryState")}="Leadership / Authority","Pitch idea, present roadmap, ask approval, send progress update, take initiative.",'
            f'IF({ref("PrimaryState")}="Wealth / Gain","Review investments, plan purchase, discuss income, ask payment, make business proposal.",'
            f'IF({ref("PrimaryState")}="Relationship / Social","Approach respectfully, send family message, start soft conversation, socialize.",'
            f'IF({ref("PrimaryState")}="Learning / Wisdom","Study, research, compare options, document, write, analyze, start course.",'
            f'IF({ref("PrimaryState")}="Execution / Problem Solving","Code, debug, run automation, validate data, fix report, handle operational work.",'
            f'IF({ref("PrimaryState")}="Travel / Movement","Leave home, start journey, visit someone, commute, attend meeting.",'
            f'IF({ref("PrimaryState")}="Purchase / Acquisition","Buy gadgets, tools, courses, software, subscriptions, equipment, useful upgrades.",'
            f'IF({ref("PrimaryState")}="Avoid / Bad Kaal","Routine work, planning, review, rest, non-critical tasks.","Daily work, admin, maintenance, light planning, routine tasks."))))))))))'
        ),
        "AvoidActions": (
            f'=IF({ref("PrimaryState")}="Golden Muhurat","Avoid wasting the window on trivial actions or impulsive decisions.",'
            f'IF({ref("PrimaryState")}="Auspicious Muhurat","Avoid reckless risk, emotional confrontation, or poorly prepared commitments.",'
            f'IF({ref("PrimaryState")}="Leadership / Authority","Avoid ego-driven message, emotional confrontation, impulsive resignation, power struggle.",'
            f'IF({ref("PrimaryState")}="Wealth / Gain","Avoid blind speculation, emotional spending, leverage, gambling-style trades.",'
            f'IF({ref("PrimaryState")}="Relationship / Social","Avoid pressure, emotional ultimatum, confrontation, forced commitment.",'
            f'IF({ref("PrimaryState")}="Learning / Wisdom","Avoid distracted study, rushed conclusions, shallow research.",'
            f'IF({ref("PrimaryState")}="Execution / Problem Solving","Avoid reckless deployment, unnecessary confrontation, emotional decision.",'
            f'IF({ref("PrimaryState")}="Travel / Movement","Avoid risky travel, angry departure, unnecessary rush.",'
            f'IF({ref("PrimaryState")}="Purchase / Acquisition","Avoid impulse buying, luxury-only spending, unresearched purchase.",'
            f'IF({ref("PrimaryState")}="Avoid / Bad Kaal","Avoid new starts, investments, commitments, sensitive messages, travel, purchases, confrontation.","Avoid major irreversible decisions unless necessary."))))))))))'
        ),
        "V05CalculationNote": '="V06 parent-state rule engine. Scores are indicator-based support signals, not guarantees."',
    }


def apply_v05_columns(ws) -> None:
    remove_old_formula_columns(ws)
    remove_existing_v05_columns(ws)
    v05_col_map = {}
    for header in V05_COLUMNS:
        v05_col_map[header] = append_or_update_column(ws, header)
    header_map = get_header_map(ws)

    for row_number in range(2, ws.max_row + 1):
        formulas = build_formulas(row_number, header_map, v05_col_map)
        for header, formula in formulas.items():
            ws.cell(row_number, v05_col_map[header], formula)


def apply_formatting(ws) -> None:
    header_map = get_header_map(ws)
    flag_fill = PatternFill("solid", fgColor="FCE4D6")
    score_fill = PatternFill("solid", fgColor="D9EAF7")
    output_fill = PatternFill("solid", fgColor="E2F0D9")

    for header in FLAG_COLUMNS:
        if header in header_map:
            cell = ws.cell(1, header_map[header])
            cell.fill = flag_fill
            cell.font = Font(bold=True)
    for header in SCORE_COLUMNS:
        if header in header_map:
            cell = ws.cell(1, header_map[header])
            cell.fill = score_fill
            cell.font = Font(bold=True)
    for header in OUTPUT_COLUMNS:
        if header in header_map:
            cell = ws.cell(1, header_map[header])
            cell.fill = output_fill
            cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        current = ws.column_dimensions[excel_col(col_idx)].width or 10
        header = str(ws.cell(1, col_idx).value or "")
        ws.column_dimensions[excel_col(col_idx)].width = max(current, min(max(len(header) + 2, 10), 28))


def choose_source_workbook() -> Path:
    return BASE_WORKBOOK


def validate_formula_text(formula: str) -> list[str]:
    issues: list[str] = []
    if not formula.startswith("="):
        issues.append("missing leading =")
    if formula.count("(") != formula.count(")"):
        issues.append("unbalanced parentheses")
    if formula.count('"') % 2 != 0:
        issues.append("odd number of double quotes")
    upper_formula = formula.upper()
    if " IN {" in upper_formula:
        issues.append("contains pseudo syntax IN {...}")
    if " THEN " in upper_formula:
        issues.append("contains pseudo syntax THEN")
    if "#REF!" in upper_formula:
        issues.append("contains #REF!")
    if "TRUE(" in upper_formula or "FALSE(" in upper_formula:
        issues.append("contains suspicious TRUE/FALSE function-like usage")
    if " TRUE" in formula or " FALSE" in formula or "(True" in formula or "(False" in formula:
        issues.append("contains Python-style True/False token")
    return issues


def validate_formula_cells(ws) -> tuple[int, list[tuple[str, str, list[str]]]]:
    validated = 0
    suspicious: list[tuple[str, str, list[str]]] = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                validated += 1
                issues = validate_formula_text(cell.value)
                if issues:
                    suspicious.append((cell.coordinate, cell.value, issues))
    return validated, suspicious


def has_duplicate_headers(ws) -> bool:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1) if ws.cell(1, col).value not in (None, "")]
    return len(headers) != len(set(headers))


def validate_output_workbook(output_path: Path) -> dict[str, object]:
    wb = load_workbook(output_path, data_only=False)
    sheetnames = wb.sheetnames
    ws = wb["EPHEMERIS_RAW"]
    header_map = get_header_map(ws)
    required_row_checks: dict[str, bool] = {}
    for header in REQUIRED_OUTPUT_HEADERS:
        col_idx = safe_col(header_map, header)
        populated = True
        for row_number in range(2, min(ws.max_row, 21) + 1):
            value = ws.cell(row_number, col_idx).value
            if value in (None, ""):
                populated = False
                break
        required_row_checks[header] = populated
    return {
        "sheetnames": sheetnames,
        "forbidden_present": [name for name in sheetnames if name in FORBIDDEN_SHEETS],
        "duplicate_headers": has_duplicate_headers(ws),
        "required_row_checks": required_row_checks,
    }


def create_minimal_debug_workbook(source_path: Path, debug_path: Path) -> None:
    wb = load_workbook(source_path)
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in {"CONFIG", "EPHEMERIS_RAW"}:
            wb.remove(wb[sheet_name])
    ws = wb["EPHEMERIS_RAW"]
    if ws.max_row > 51:
        ws.delete_rows(52, ws.max_row - 51)
    if hasattr(ws.conditional_formatting, "_cf_rules"):
        ws.conditional_formatting._cf_rules.clear()
    for row in ws.iter_rows():
        for cell in row:
            if cell.has_style:
                cell._style = copy(cell._style)
    wb.save(debug_path)


def main() -> None:
    source_path = choose_source_workbook()
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(source_path)
    if "INDICATORS" in wb.sheetnames:
        raise ValueError("INDICATORS sheet is not allowed in the V06 workbook.")
    for sheet_name in wb.sheetnames:
        clear_conditional_formatting(wb[sheet_name])
    ws = wb["EPHEMERIS_RAW"]
    apply_v05_columns(ws)
    apply_formatting(ws)
    validated_count, suspicious = validate_formula_cells(ws)
    for coord, formula, issues in suspicious[:MAX_SUSPICIOUS_PRINT]:
        print(f"SUSPICIOUS {coord}: {'; '.join(issues)} :: {formula}")
    if len(suspicious) > MAX_SUSPICIOUS_PRINT:
        print(f"SUSPICIOUS_TRUNCATED={len(suspicious) - MAX_SUSPICIOUS_PRINT}")
    wb.save(OUTPUT_WORKBOOK)
    validation = validate_output_workbook(OUTPUT_WORKBOOK)
    if validation["forbidden_present"] or validation["duplicate_headers"] or not all(validation["required_row_checks"].values()):
        debug_path = DEBUG_DIR / "MuhuratFinder_V06_ParentStateEngine_DEBUG.xlsx"
        create_minimal_debug_workbook(OUTPUT_WORKBOOK, debug_path)
        print(f"DEBUG_WORKBOOK={debug_path}")
    print(f"OUTPUT={OUTPUT_WORKBOOK}")
    print(f"FORMULA_CELLS_VALIDATED={validated_count}")
    print(f"SUSPICIOUS_FORMULAS={len(suspicious)}")
    print(f"VALIDATION={validation}")


if __name__ == "__main__":
    main()
