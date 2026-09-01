"""V03 raw-only workbook and CSV export."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

import pandas as pd
import swisseph as swe
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .astronomy import SwissEphemerisEngine, local_datetime_to_jd, normalize_degrees
from .config import (
    DATA_DIR,
    DEFAULT_LOCATION,
    DEFAULT_RAW_JSON_PATH,
    NAKSHATRA_NAMES,
    SIGN_NAMES,
    V03_STATUS_PATH,
)
from .rules import build_abhijit, build_choghadiya, build_horas, build_rahu_yamaganda


BUILD_DIR = Path(__file__).resolve().parent.parent
V03_DIR = BUILD_DIR.parent
OUTPUT_XLSX = V03_DIR / "MuhuratFinder_V03.xlsx"
EXPORT_DIR = BUILD_DIR / "data" / "v03_exports"
IST = "Asia/Kolkata"
PLANETS = ["Sun", "Moon", "Mars", "Mercury", "Jupiter", "Venus", "Saturn", "Rahu", "Ketu"]

CONFIG_CELL_MAP = {
    "from_date": "B3",
    "end_date": "B4",
    "ayanamsa": "B5",
    "timezone": "B6",
    "dst": "B7",
    "status": "E2",
}

RAW_HEADERS = [
    "Date",
    "Day",
    "Start",
    "End",
    "Min",
    "StartDateTime",
    "EndDateTime",
    "Sunrise",
    "Sunset",
    "Timezone",
    "DST",
    "Ayanamsa",
    "AyanamsaDeg",
    "Paksha",
    "Tithi",
    "TithiNo",
    "MoonNakshatra",
    "MoonPada",
    "LagnaNakshatra",
    "LagnaPada",
    "Yoga",
    "Karana",
    "Choghadiya",
    "Hora",
    "Abhijit",
    "RahuKaal",
    "Yamaganda",
    "Gulika",
    "Durmuhurta",
    "Varjyam",
    "TithiKshaya",
    "NakshatraKshaya",
    "LagnaSign",
    "LagnaDeg",
    "MoonSign",
    "MoonDeg",
    "MoonHouse",
    "MoonRetro",
]
for _planet in [planet for planet in PLANETS if planet != "Moon"]:
    RAW_HEADERS.extend(
        [
            f"{_planet}Sign",
            f"{_planet}Deg",
            f"{_planet}House",
            f"{_planet}Nakshatra",
            f"{_planet}Pada",
            f"{_planet}Retro",
        ]
    )


@dataclass(frozen=True)
class RawConfig:
    """Minimal V03 workbook configuration."""

    from_date: date
    end_date: date
    ayanamsa: str = "Lahiri"
    timezone: str = IST
    dst: str = "N"


def default_raw_config() -> RawConfig:
    """Return default V03 config."""

    return RawConfig(
        from_date=date(2026, 4, 28),
        end_date=date(2026, 5, 31),
    )


def read_config_from_workbook(path: Path) -> RawConfig:
    """Read minimal V03 config from workbook."""

    wb = load_workbook(path, data_only=True)
    ws = wb["CONFIG"]
    defaults = default_raw_config()

    def as_date(value: Any, fallback: date) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str) and value.strip():
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    pass
        return fallback

    return RawConfig(
        from_date=as_date(ws[CONFIG_CELL_MAP["from_date"]].value, defaults.from_date),
        end_date=as_date(ws[CONFIG_CELL_MAP["end_date"]].value, defaults.end_date),
        ayanamsa=str(ws[CONFIG_CELL_MAP["ayanamsa"]].value or defaults.ayanamsa),
        timezone=str(ws[CONFIG_CELL_MAP["timezone"]].value or defaults.timezone),
        dst=str(ws[CONFIG_CELL_MAP["dst"]].value or defaults.dst),
    )


def floor_minute(timestamp: pd.Timestamp) -> pd.Timestamp:
    """Floor timestamp to minute precision in IST."""

    return timestamp.tz_convert(IST).floor("min")


def sign_count(from_sign: str, to_sign: str) -> int:
    """Return 1-based zodiac count from one sign to another."""

    return ((SIGN_NAMES.index(to_sign) - SIGN_NAMES.index(from_sign)) % 12) + 1


def whole_sign_house(lagna_sign: str, planet_sign: str) -> int:
    """Return whole-sign house of a planet from Lagna."""

    return sign_count(lagna_sign, planet_sign)


def longitude_parts(longitude: float) -> Dict[str, Any]:
    """Return sign, degree, nakshatra, and pada for a sidereal longitude."""

    lon = normalize_degrees(longitude)
    sign_index = int(lon // 30)
    degree_in_sign = lon % 30
    nak_span = 360.0 / len(NAKSHATRA_NAMES)
    nak_index = min(int(lon // nak_span), len(NAKSHATRA_NAMES) - 1)
    pada = int(((lon % nak_span) / (nak_span / 4.0))) + 1
    return {
        "longitude": round(lon, 6),
        "sign": SIGN_NAMES[sign_index],
        "degree": round(degree_in_sign, 4),
        "nakshatra": NAKSHATRA_NAMES[nak_index],
        "pada": min(pada, 4),
    }


TITHI_NAMES = [
    "Pratipada",
    "Dwitiya",
    "Tritiya",
    "Chaturthi",
    "Panchami",
    "Shashthi",
    "Saptami",
    "Ashtami",
    "Navami",
    "Dashami",
    "Ekadashi",
    "Dwadashi",
    "Trayodashi",
    "Chaturdashi",
    "Purnima",
]

YOGA_NAMES = [
    "Vishkambha", "Priti", "Ayushman", "Saubhagya", "Shobhana", "Atiganda", "Sukarma", "Dhriti", "Shoola",
    "Ganda", "Vriddhi", "Dhruva", "Vyaghata", "Harshana", "Vajra", "Siddhi", "Vyatipata", "Variyana", "Parigha",
    "Shiva", "Siddha", "Sadhya", "Shubha", "Shukla", "Brahma", "Indra", "Vaidhriti",
]


def tithi_details(sun_longitude: float, moon_longitude: float) -> Dict[str, Any]:
    """Compute raw tithi and paksha."""

    diff = normalize_degrees(moon_longitude - sun_longitude)
    tithi_number = int(diff // 12.0) + 1
    paksha = "Shukla Paksha" if diff < 180.0 else "Krishna Paksha"
    paksha_tithi = ((tithi_number - 1) % 15) + 1
    if paksha_tithi == 15:
        name = "Purnima" if paksha == "Shukla Paksha" else "Amavasya"
    else:
        name = TITHI_NAMES[paksha_tithi - 1]
    return {
        "paksha": paksha,
        "tithi_number": tithi_number,
        "tithi": name,
    }


def yoga_name(sun_longitude: float, moon_longitude: float) -> str:
    """Compute nitya yoga from Sun and Moon longitudes."""

    index = int(normalize_degrees(sun_longitude + moon_longitude) // (360.0 / 27.0))
    return YOGA_NAMES[min(index, 26)]


def karana_name(sun_longitude: float, moon_longitude: float) -> str:
    """Compute karana from the half-tithi index."""

    diff = normalize_degrees(moon_longitude - sun_longitude)
    half_index = int(diff // 6.0) + 1
    if half_index == 1:
        return "Kimstughna"
    if half_index >= 58:
        return ["Shakuni", "Chatushpada", "Naga"][min(half_index - 58, 2)]
    repeating = ["Bava", "Balava", "Kaulava", "Taitila", "Gara", "Vanija", "Vishti"]
    return repeating[(half_index - 2) % 7]


def find_segment(moment: datetime, segments: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    """Find segment containing a moment."""

    for segment in segments:
        if segment["start"] <= moment < segment["end"]:
            return segment
    return segments[-1]


def overlaps(start: datetime, end: datetime, interval: Dict[str, datetime]) -> bool:
    """Check overlap with a half-open interval."""

    return start < interval["end"] and end > interval["start"]


def build_day_payload(local_day: date, astro: SwissEphemerisEngine) -> Dict[str, Any]:
    """Build only the raw interval inputs for one day."""

    astro_day = astro.sunrise_sunset(local_day)
    horas = build_horas(local_day, astro_day["sunrise"], astro_day["sunset"], astro_day["next_sunrise"])
    choghadiya = build_choghadiya(local_day, astro_day["sunrise"], astro_day["sunset"], astro_day["next_sunrise"])
    doshas = build_rahu_yamaganda(local_day, astro_day["sunrise"], astro_day["sunset"])
    abhijit = build_abhijit(astro_day["solar_noon"])
    return {
        "date": local_day.isoformat(),
        "weekday": local_day.strftime("%A"),
        "sunrise": astro_day["sunrise"].isoformat(),
        "sunset": astro_day["sunset"].isoformat(),
        "next_sunrise": astro_day["next_sunrise"].isoformat(),
        "solar_noon": astro_day["solar_noon"].isoformat(),
        "horas": horas,
        "choghadiya": choghadiya,
        "rahu_kaal": doshas["rahu_kaal"],
        "yamaganda": doshas["yamaganda"],
        "abhijit_muhurat": abhijit,
    }


def build_raw_rows(config: RawConfig) -> List[Dict[str, Any]]:
    """Build raw-only EPHEMERIS_RAW rows."""

    if config.end_date < config.from_date:
        raise ValueError("End Date must be on or after From Date.")

    astro = SwissEphemerisEngine(DEFAULT_LOCATION)
    rows: List[Dict[str, Any]] = []
    previous_signature: Tuple[Any, ...] | None = None
    final_next_sunrise: datetime | None = None

    for offset in range((config.end_date - config.from_date).days + 1):
        local_day = config.from_date + timedelta(days=offset)
        day_record = build_day_payload(local_day, astro)
        horas = day_record["horas"]
        choghadiya = day_record["choghadiya"]
        rahu_kaal = day_record["rahu_kaal"]
        yamaganda = day_record["yamaganda"]
        abhijit = day_record["abhijit_muhurat"]
        sunrise = floor_minute(pd.Timestamp(day_record["sunrise"]))
        sunset = floor_minute(pd.Timestamp(day_record["sunset"]))
        next_sunrise = floor_minute(pd.Timestamp(day_record["next_sunrise"]))
        final_next_sunrise = next_sunrise.to_pydatetime().replace(tzinfo=None)
        minute_ts = sunrise

        while minute_ts < next_sunrise:
            start_dt = minute_ts.to_pydatetime()
            end_guess = (minute_ts + timedelta(minutes=1)).to_pydatetime()
            snapshot = astro.planetary_snapshot(start_dt)
            panchang = astro.panchang_snapshot(start_dt)
            planets = snapshot["planets"]
            sun_lon = float(planets["Sun"]["longitude"])
            moon_lon = float(planets["Moon"]["longitude"])
            moon_parts = longitude_parts(moon_lon)
            lagna_parts = longitude_parts(float(panchang["ascendant_longitude"]))
            tithi = tithi_details(sun_lon, moon_lon)
            yoga = yoga_name(sun_lon, moon_lon)
            karana = karana_name(sun_lon, moon_lon)
            jd = local_datetime_to_jd(start_dt.replace(tzinfo=minute_ts.tzinfo))
            ayanamsa_value = round(swe.get_ayanamsa_ut(jd), 6)
            hora_segment = find_segment(start_dt, horas)
            choghadiya_segment = find_segment(start_dt, choghadiya)

            row: Dict[str, Any] = {
                "Date": start_dt.date(),
                "Day": str(day_record["weekday"]),
                "Start": start_dt.time(),
                "End": None,
                "Min": None,
                "StartDateTime": start_dt.replace(tzinfo=None),
                "EndDateTime": None,
                "Sunrise": sunrise.to_pydatetime().replace(tzinfo=None),
                "Sunset": sunset.to_pydatetime().replace(tzinfo=None),
                "Timezone": config.timezone,
                "DST": config.dst,
                "Ayanamsa": config.ayanamsa,
                "AyanamsaDeg": ayanamsa_value,
                "Paksha": tithi["paksha"],
                "Tithi": tithi["tithi"],
                "TithiNo": tithi["tithi_number"],
                "MoonNakshatra": moon_parts["nakshatra"],
                "MoonPada": moon_parts["pada"],
                "LagnaNakshatra": lagna_parts["nakshatra"],
                "LagnaPada": lagna_parts["pada"],
                "Yoga": yoga,
                "Karana": karana,
                "Choghadiya": str(choghadiya_segment["type"]),
                "Hora": str(hora_segment["ruler"]),
                "Abhijit": "Y" if overlaps(start_dt, end_guess, abhijit) else "N",
                "RahuKaal": "Y" if overlaps(start_dt, end_guess, rahu_kaal) else "N",
                "Yamaganda": "Y" if overlaps(start_dt, end_guess, yamaganda) else "N",
                "Gulika": "N",
                "Durmuhurta": "N",
                "Varjyam": "N",
                "TithiKshaya": "N",
                "NakshatraKshaya": "N",
                "LagnaSign": lagna_parts["sign"],
                "LagnaDeg": lagna_parts["degree"],
                "MoonSign": moon_parts["sign"],
                "MoonDeg": moon_parts["degree"],
                "MoonHouse": whole_sign_house(lagna_parts["sign"], moon_parts["sign"]),
                "MoonRetro": "Y" if planets["Moon"].get("retrograde") else "N",
            }

            for planet in [planet for planet in PLANETS if planet != "Moon"]:
                parts = longitude_parts(float(planets[planet]["longitude"]))
                row[f"{planet}Sign"] = parts["sign"]
                row[f"{planet}Deg"] = parts["degree"]
                row[f"{planet}House"] = whole_sign_house(lagna_parts["sign"], parts["sign"])
                row[f"{planet}Nakshatra"] = parts["nakshatra"]
                row[f"{planet}Pada"] = parts["pada"]
                row[f"{planet}Retro"] = "Y" if planets[planet].get("retrograde") else "N"

            # Keep exact degrees at the interval start, but do not let smoothly
            # changing degree values force one row per minute. A new row should
            # appear only when the raw state changes meaningfully.
            signature_fields = [
                header
                for header in RAW_HEADERS
                if header
                not in {
                    "Date",
                    "Start",
                    "End",
                    "Min",
                    "StartDateTime",
                    "EndDateTime",
                    "Sunrise",
                    "Sunset",
                    "AyanamsaDeg",
                }
                and not header.endswith("Deg")
            ]
            signature = tuple(row.get(field) for field in signature_fields)
            if signature == previous_signature:
                minute_ts += timedelta(minutes=1)
                continue

            previous_signature = signature
            rows.append(row)
            minute_ts += timedelta(minutes=1)

    for index, row in enumerate(rows):
        if index + 1 < len(rows):
            end_dt = rows[index + 1]["StartDateTime"]
        else:
            end_dt = final_next_sunrise or datetime.combine(config.end_date + timedelta(days=1), datetime.min.time())
        row["EndDateTime"] = end_dt
        row["End"] = end_dt.time()
        row["Min"] = max(0, round((end_dt - row["StartDateTime"]).total_seconds() / 60.0))

    return rows


def write_sheet(ws, rows: Sequence[Dict[str, Any]], headers: Sequence[str]) -> None:
    """Write dictionaries to worksheet with basic styling."""

    ws.delete_rows(1, ws.max_row)
    ws.append(list(headers))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(header) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        width = min(max(len(str(ws.cell(1, column).value or "")) + 2, 10), 24)
        ws.column_dimensions[letter].width = width


def build_workbook(config: RawConfig, output_path: Path = OUTPUT_XLSX) -> Path:
    """Generate the V03 raw-only workbook."""

    rows = build_raw_rows(config)
    payload = {
        "generated_at": datetime.now().isoformat(),
        "location": {
            "name": DEFAULT_LOCATION.name,
            "latitude": DEFAULT_LOCATION.latitude,
            "longitude": DEFAULT_LOCATION.longitude,
            "timezone": DEFAULT_LOCATION.timezone,
            "altitude_m": DEFAULT_LOCATION.altitude_m,
        },
        "from_date": config.from_date.isoformat(),
        "end_date": config.end_date.isoformat(),
        "row_count": len(rows),
    }
    DEFAULT_RAW_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    DEFAULT_RAW_JSON_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    wb = Workbook()
    config_ws = wb.active
    config_ws.title = "CONFIG"
    config_ws["A1"] = "Muhurat Finder V03"
    config_ws["A2"] = "Raw-only Python export. Excel should own all formulas, scoring, purpose logic, and recommendations."
    config_ws["A3"] = "From Date"
    config_ws[CONFIG_CELL_MAP["from_date"]] = config.from_date
    config_ws["A4"] = "End Date"
    config_ws[CONFIG_CELL_MAP["end_date"]] = config.end_date
    config_ws["A5"] = "Ayanamsa"
    config_ws[CONFIG_CELL_MAP["ayanamsa"]] = config.ayanamsa
    config_ws["A6"] = "Timezone"
    config_ws[CONFIG_CELL_MAP["timezone"]] = config.timezone
    config_ws["A7"] = "DST"
    config_ws[CONFIG_CELL_MAP["dst"]] = config.dst
    config_ws["E2"] = "Status"
    config_ws[CONFIG_CELL_MAP["status"]] = f"Built {len(rows)} raw rows"
    config_ws.column_dimensions["A"].width = 18
    config_ws.column_dimensions["B"].width = 18
    config_ws.column_dimensions["E"].width = 12
    config_ws.column_dimensions["F"].width = 26
    config_ws[CONFIG_CELL_MAP["from_date"]].number_format = "dd-mmm-yyyy"
    config_ws[CONFIG_CELL_MAP["end_date"]].number_format = "dd-mmm-yyyy"

    raw_ws = wb.create_sheet("EPHEMERIS_RAW")
    write_sheet(raw_ws, rows, RAW_HEADERS)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_csv_exports(config: RawConfig, csv_dir: Path = EXPORT_DIR) -> Path:
    """Write only EPHEMERIS_RAW.csv for V03."""

    rows = build_raw_rows(config)
    csv_dir.mkdir(parents=True, exist_ok=True)
    V03_STATUS_PATH.parent.mkdir(parents=True, exist_ok=True)
    path = csv_dir / "EPHEMERIS_RAW.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=RAW_HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    V03_STATUS_PATH.write_text(
        json.dumps(
            {
                "ok": True,
                "from_date": config.from_date.isoformat(),
                "end_date": config.end_date.isoformat(),
                "rows": len(rows),
                "export_path": str(path),
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    return path
