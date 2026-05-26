"""Export V05 workbook rows into static JSON files for the V06 dashboard."""

from __future__ import annotations

import json
import os
import argparse
import platform
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    import win32com.client  # type: ignore
except Exception:  # pragma: no cover - optional dependency at runtime
    win32com = None
else:  # pragma: no cover - import alias for lint clarity
    win32com = win32com.client

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    from backports.zoneinfo import ZoneInfo  # type: ignore


ROOT_DIR = Path(__file__).resolve().parent
DATA_SOURCE_DIR = ROOT_DIR / "data_source"
DEFAULT_SOURCE_WORKBOOK = DATA_SOURCE_DIR / "MuhuratFinder_V05_ParentStateEngine.xlsx"
PUBLIC_DATA_DIR = ROOT_DIR / "web" / "public" / "data"

WINDOWS_JSON = PUBLIC_DATA_DIR / "windows.json"
DAY_SUMMARY_JSON = PUBLIC_DATA_DIR / "day_summary.json"
CONFIG_JSON = PUBLIC_DATA_DIR / "config.json"
MUHURAT_DATA_JSON = PUBLIC_DATA_DIR / "muhurat-data.json"

AVAILABLE_CATEGORIES = [
    "overall",
    "golden",
    "auspicious",
    "leadership",
    "wealth",
    "relationship",
    "learning",
    "execution",
    "travel",
    "purchase",
    "avoid",
]

CRITICAL_HEADERS = [
    "Date",
    "Day",
    "Start",
    "End",
    "StartDateTime",
    "EndDateTime",
    "Sunrise",
    "Sunset",
    "Timezone",
]

OPTIONAL_HEADERS = [
    "Paksha",
    "Tithi",
    "TithiNo",
    "MoonNakshatra",
    "MoonPada",
    "Yoga",
    "Karana",
    "Choghadiya",
    "Hora",
    "Abhijit",
    "RahuKaal",
    "Yamaganda",
    "Gulika",
    "Durmuhurta",
    "Varjyam",
    "LagnaSign",
    "LagnaDeg",
    "LagnaNakshatra",
    "LagnaPada",
    "MoonSign",
    "MoonDeg",
    "MoonHouse",
    "EventLocationName",
    "EventLatitude",
    "EventLongitude",
    "EventTimezone",
    "EventDST",
    "NatalMoonSign",
    "NatalNakshatra",
    "NatalLagna",
    "PrimaryState",
    "PrimaryStateReason",
    "SecondaryStates",
    "SecondaryStateReason",
    "RiskLevel",
    "RiskReason",
    "BestActions",
    "AvoidActions",
    "AvoidScore",
    "GoldenScore",
    "AuspiciousScore",
    "LeadershipScore",
    "WealthScore",
    "RelationshipScore",
    "LearningScore",
    "ExecutionScore",
    "TravelScore",
    "PurchaseScore",
]

SCORE_HEADERS = {
    "avoid": "AvoidScore",
    "golden": "GoldenScore",
    "auspicious": "AuspiciousScore",
    "leadership": "LeadershipScore",
    "wealth": "WealthScore",
    "relationship": "RelationshipScore",
    "learning": "LearningScore",
    "execution": "ExecutionScore",
    "travel": "TravelScore",
    "purchase": "PurchaseScore",
}

V05_REQUIRED_HEADERS = [
    "PrimaryState",
    "PrimaryStateReason",
    "SecondaryStates",
    "RiskLevel",
    "RiskReason",
    "BestActions",
    "AvoidActions",
    "AvoidScore",
    "GoldenScore",
    "AuspiciousScore",
    "LeadershipScore",
    "WealthScore",
    "RelationshipScore",
    "LearningScore",
    "ExecutionScore",
    "TravelScore",
    "PurchaseScore",
]

DEBUG_SAMPLE_HEADERS = [
    "Date",
    "Start",
    "End",
    "PrimaryState",
    "RiskLevel",
    "AvoidScore",
    "GoldenScore",
    "AuspiciousScore",
    "LeadershipScore",
    "WealthScore",
    "RelationshipScore",
    "LearningScore",
    "ExecutionScore",
    "TravelScore",
    "PurchaseScore",
]


@dataclass
class ExportStats:
    source_workbook: Path
    raw_rows_read: int
    windows_count: int
    day_count: int
    first_exported_date: str | None
    last_exported_date: str | None
    unique_date_count: int
    first_sample_dates: list[str]
    last_sample_dates: list[str]
    missing_value_count: int
    missing_score_count: int
    score_columns_found: bool
    formula_cached_values_missing: bool
    detected_headers: list[str]
    missing_v05_headers: list[str]
    debug_samples: list[dict[str, Any]]
    primary_state_present_count: int
    all_scores_present_count: int
    excel_recalculation_ran: bool
    excel_recalculation_warning: str | None
    blocked_error: str | None
    warnings: list[str]


def display_path(path: Path) -> str:
    try:
        return os.path.relpath(path, ROOT_DIR)
    except ValueError:
        return str(path)


def choose_source_workbook(source_override: str | None = None) -> Path:
    source_workbook = Path(source_override) if source_override else DEFAULT_SOURCE_WORKBOOK
    if not source_workbook.is_absolute():
        source_workbook = (Path.cwd() / source_workbook).resolve() if source_override else source_workbook.resolve()
    if not source_workbook.exists():
        raise FileNotFoundError(f"Source workbook not found: {source_workbook}")
    try:
        load_workbook(source_workbook, read_only=True, data_only=False).close()
    except Exception as exc:
        raise ValueError(f"Source workbook is not a valid readable workbook: {source_workbook}") from exc
    return source_workbook


def calculate_workbook_with_excel_com(source_path: Path) -> None:
    if win32com is None:
        raise RuntimeError("win32com.client is not available.")
    excel = None
    workbook = None
    try:
        excel = win32com.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(source_path))
        excel.CalculateFullRebuild()
        workbook.Save()
        workbook.Close(SaveChanges=True)
        workbook = None
    finally:  # pragma: no cover - depends on local Excel
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=True)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def calculate_workbook_with_powershell_com(source_path: Path) -> None:
    script = f"""
$excel = $null
$workbook = $null
try {{
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $workbook = $excel.Workbooks.Open('{str(source_path).replace("'", "''")}')
  $excel.CalculateFullRebuild()
  $workbook.Save()
  $workbook.Close($true)
}} finally {{
  if ($workbook -ne $null) {{
    try {{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null }} catch {{ }}
  }}
  if ($excel -ne $null) {{
    try {{ $excel.Quit() }} catch {{ }}
    try {{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null }} catch {{ }}
  }}
  [GC]::Collect()
  [GC]::WaitForPendingFinalizers()
}}
"""
    result = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        check=False,
        cwd=ROOT_DIR,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        details = [
            line.strip()
            for line in (result.stderr or result.stdout or "").splitlines()
            if line.strip()
        ]
        raise RuntimeError(details[-1] if details else f"PowerShell exited with code {result.returncode}.")


def maybe_recalculate_workbook(source_workbook: Path, should_recalculate: bool) -> tuple[bool, str | None]:
    if not should_recalculate:
        return False, None
    if platform.system() != "Windows":
        return False, "Excel COM recalculation skipped: not running on Windows."
    if win32com is not None:
        calculate_workbook_with_excel_com(source_workbook)
        return True, None
    try:
        calculate_workbook_with_powershell_com(source_workbook)
        return True, "win32com.client is not available; used PowerShell Excel COM fallback."
    except Exception as exc:
        detail = str(exc).splitlines()[-1] if str(exc) else "unknown error"
        return False, f"Excel COM recalculation failed or is unavailable: {detail}"


def header_map(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, column).value): column
        for column in range(1, ws.max_column + 1)
        if ws.cell(1, column).value not in (None, "")
    }


def require_headers(headers: dict[str, int], names: list[str]) -> None:
    missing = [name for name in names if name not in headers]
    if missing:
        raise ValueError(f"Missing critical EPHEMERIS_RAW columns: {', '.join(missing)}")


def excel_value(ws, row_number: int, headers: dict[str, int], header_name: str) -> Any:
    column = headers.get(header_name)
    if column is None:
        return None
    return ws.cell(row_number, column).value


def to_local_datetime(value: Any, timezone_name: str | None) -> datetime | None:
    if value in (None, ""):
        return None
    tz_name = timezone_name or "Asia/Kolkata"
    tz = ZoneInfo(tz_name)
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=tz)
        return value.astimezone(tz)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=tz)
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
        try:
            parsed = datetime.fromisoformat(candidate)
        except ValueError:
            try:
                parsed = datetime.strptime(candidate, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    parsed = datetime.strptime(candidate, "%Y-%m-%d %H:%M")
                except ValueError:
                    return None
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=tz)
        return parsed.astimezone(tz)
    return None


def to_iso(value: Any, timezone_name: str | None) -> str | None:
    local_dt = to_local_datetime(value, timezone_name)
    return local_dt.isoformat() if local_dt else None


def to_date_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def to_time_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    text = str(value).strip()
    if not text:
        return None
    if len(text) == 5:
        return f"{text}:00"
    return text


def to_short_time(iso_value: str | None) -> str | None:
    if not iso_value:
        return None
    return datetime.fromisoformat(iso_value).strftime("%H:%M")


def shifted_datetime_to_date(value: Any, target_date: date, timezone_name: str | None) -> datetime | None:
    local_dt = to_local_datetime(value, timezone_name)
    if local_dt is None:
        return None
    return datetime.combine(
        target_date,
        time(local_dt.hour, local_dt.minute, local_dt.second, local_dt.microsecond),
        tzinfo=local_dt.tzinfo,
    )


def to_number(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return int(numeric) if numeric.is_integer() else round(numeric, 4)


def normalize_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value).strip()


def get_row_scores(row: dict[str, Any]) -> dict[str, int | None]:
    scores: dict[str, int | None] = {}
    for key, header in SCORE_HEADERS.items():
        numeric = to_number(row.get(header))
        scores[key] = int(numeric) if numeric is not None else None
    return scores


def first_present(windows: list[dict[str, Any]], field: str) -> Any:
    for window in windows:
        value = window.get(field)
        if value not in (None, ""):
            return value
    return None


def read_config_lookup(wb) -> dict[str, Any]:
    if "CONFIG" not in wb.sheetnames:
        return {}
    ws = wb["CONFIG"]
    config: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        for index, cell in enumerate(cells[:-1]):
            if cell not in (None, "") and cells[index + 1] not in (None, ""):
                config[str(cell)] = cells[index + 1]
    return config


def debug_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def overall_score(scores: dict[str, int | None]) -> int | None:
    avoid = scores.get("avoid")
    if avoid is not None and avoid >= 80:
        return 0
    positives = [
        scores.get(key)
        for key in AVAILABLE_CATEGORIES
        if key not in {"overall", "avoid"} and scores.get(key) is not None
    ]
    return max(positives) if positives else None


def export_windows_and_summary(source_workbook: Path, should_recalculate: bool = True) -> ExportStats:
    warnings: list[str] = []
    excel_recalculation_ran, excel_recalculation_warning = maybe_recalculate_workbook(source_workbook, should_recalculate)
    if excel_recalculation_warning:
        warnings.append(excel_recalculation_warning)

    wb = load_workbook(source_workbook, data_only=True, read_only=True)
    formula_wb = load_workbook(source_workbook, data_only=False, read_only=True)
    ws = wb["EPHEMERIS_RAW"]
    formula_ws = formula_wb["EPHEMERIS_RAW"]
    config_lookup = read_config_lookup(wb)
    row_iter = ws.iter_rows(values_only=True)
    formula_row_iter = formula_ws.iter_rows(values_only=True)
    raw_headers = next(row_iter)
    next(formula_row_iter)
    headers = {
        str(value): index + 1
        for index, value in enumerate(raw_headers)
        if value not in (None, "")
    }
    detected_headers = [str(value) for value in raw_headers if value not in (None, "")]
    missing_v05_headers = [name for name in V05_REQUIRED_HEADERS if name not in headers]
    score_columns_found = all(header in headers for header in SCORE_HEADERS.values())
    if missing_v05_headers:
        raise ValueError(
            "Wrong source workbook. V05 parent-state columns not found. Copy the valid V05 workbook into v06_dashboard/data_source/."
        )
    require_headers(headers, CRITICAL_HEADERS)

    PUBLIC_DATA_DIR.mkdir(parents=True, exist_ok=True)

    windows: list[dict[str, Any]] = []
    missing_values = 0
    missing_score_count = 0
    formula_cached_values_missing = False
    debug_samples: list[dict[str, Any]] = []
    raw_rows_read = 0
    for row_values in row_iter:
        formula_row_values = next(formula_row_iter, ())
        raw_rows_read += 1
        start_dt_raw = row_values[headers["StartDateTime"] - 1]
        end_dt_raw = row_values[headers["EndDateTime"] - 1]
        timezone_name = normalize_text(row_values[headers.get("EventTimezone", headers["Timezone"]) - 1]) or normalize_text(row_values[headers["Timezone"] - 1]) or "Asia/Kolkata"

        if start_dt_raw in (None, "") and end_dt_raw in (None, ""):
            continue

        raw_row = {
            name: row_values[headers[name] - 1] if name in headers else None
            for name in (CRITICAL_HEADERS + OPTIONAL_HEADERS)
        }

        window = {
            "date": to_date_text(raw_row.get("Date")),
            "day": normalize_text(raw_row.get("Day")),
            "start": to_time_text(raw_row.get("Start")),
            "end": to_time_text(raw_row.get("End")),
            "startDateTime": to_iso(raw_row.get("StartDateTime"), timezone_name),
            "endDateTime": to_iso(raw_row.get("EndDateTime"), timezone_name),
            "sunrise": to_iso(raw_row.get("Sunrise"), timezone_name),
            "sunset": to_iso(raw_row.get("Sunset"), timezone_name),
            "timezone": normalize_text(raw_row.get("Timezone")) or timezone_name,
            "paksha": normalize_text(raw_row.get("Paksha")),
            "tithi": normalize_text(raw_row.get("Tithi")),
            "tithiNo": to_number(raw_row.get("TithiNo")),
            "moonNakshatra": normalize_text(raw_row.get("MoonNakshatra")),
            "moonPada": to_number(raw_row.get("MoonPada")),
            "moonSign": normalize_text(raw_row.get("MoonSign")),
            "moonDeg": to_number(raw_row.get("MoonDeg")),
            "moonHouse": to_number(raw_row.get("MoonHouse")),
            "lagnaSign": normalize_text(raw_row.get("LagnaSign")),
            "lagnaDeg": to_number(raw_row.get("LagnaDeg")),
            "lagnaNakshatra": normalize_text(raw_row.get("LagnaNakshatra")),
            "lagnaPada": to_number(raw_row.get("LagnaPada")),
            "yoga": normalize_text(raw_row.get("Yoga")),
            "karana": normalize_text(raw_row.get("Karana")),
            "choghadiya": normalize_text(raw_row.get("Choghadiya")),
            "hora": normalize_text(raw_row.get("Hora")),
            "abhijit": normalize_text(raw_row.get("Abhijit")),
            "rahuKaal": normalize_text(raw_row.get("RahuKaal")),
            "yamaganda": normalize_text(raw_row.get("Yamaganda")),
            "gulika": normalize_text(raw_row.get("Gulika")),
            "durmuhurta": normalize_text(raw_row.get("Durmuhurta")),
            "varjyam": normalize_text(raw_row.get("Varjyam")),
            "eventLocationName": normalize_text(raw_row.get("EventLocationName")),
            "eventLatitude": to_number(raw_row.get("EventLatitude")),
            "eventLongitude": to_number(raw_row.get("EventLongitude")),
            "eventTimezone": normalize_text(raw_row.get("EventTimezone")),
            "eventDST": normalize_text(raw_row.get("EventDST")),
            "natalMoonSign": normalize_text(raw_row.get("NatalMoonSign")),
            "natalNakshatra": normalize_text(raw_row.get("NatalNakshatra")),
            "natalLagna": normalize_text(raw_row.get("NatalLagna")),
            "primaryState": normalize_text(raw_row.get("PrimaryState")),
            "primaryStateReason": normalize_text(raw_row.get("PrimaryStateReason")),
            "secondaryStates": normalize_text(raw_row.get("SecondaryStates")),
            "secondaryStateReason": normalize_text(raw_row.get("SecondaryStateReason")),
            "riskLevel": normalize_text(raw_row.get("RiskLevel")),
            "riskReason": normalize_text(raw_row.get("RiskReason")),
            "bestActions": normalize_text(raw_row.get("BestActions")),
            "avoidActions": normalize_text(raw_row.get("AvoidActions")),
        }
        scores = get_row_scores(raw_row)
        window["scores"] = scores

        for key, header in SCORE_HEADERS.items():
            if scores.get(key) is not None or header not in headers:
                continue
            missing_score_count += 1
            formula_value = formula_row_values[headers[header] - 1] if len(formula_row_values) >= headers[header] else None
            if isinstance(formula_value, str) and formula_value.startswith("="):
                formula_cached_values_missing = True

        if len(debug_samples) < 20:
            debug_samples.append(
                {
                    name: debug_value(raw_row.get(name))
                    for name in DEBUG_SAMPLE_HEADERS
                }
            )

        if any(window[field] is None for field in ["primaryState", "riskLevel", "bestActions", "avoidActions"]):
            missing_values += 1
        windows.append(window)

    wb.close()
    formula_wb.close()

    if not windows:
        raise ValueError("No EPHEMERIS_RAW rows were exported.")

    windows.sort(key=lambda item: item["startDateTime"] or "")
    primary_state_present_count = sum(1 for item in windows if item.get("primaryState"))
    all_scores_present_count = sum(
        1
        for item in windows
        if all(item.get("scores", {}).get(key) is not None for key in SCORE_HEADERS)
    )

    day_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for window in windows:
        if window["date"]:
            day_groups[window["date"]].append(window)

    day_summaries: list[dict[str, Any]] = []
    for day_key in sorted(day_groups):
        items = day_groups[day_key]
        sample = items[0]
        timezone_name = sample.get("eventTimezone") or sample.get("timezone") or "Asia/Kolkata"
        target_date = date.fromisoformat(day_key)
        sunrise_item = next(
            (
                item
                for item in items
                if (to_date_text(to_local_datetime(item.get("sunrise"), timezone_name)) == day_key)
            ),
            None,
        )
        sunrise_source = sunrise_item or sample
        sunrise_dt = shifted_datetime_to_date(sunrise_source.get("sunrise"), target_date, timezone_name)
        if sunrise_dt is None:
            sunrise_dt = datetime.combine(target_date, time.min, tzinfo=ZoneInfo(timezone_name))
        midnight_dt = datetime.combine(sunrise_dt.date() + timedelta(days=1), time.min, tzinfo=sunrise_dt.tzinfo)
        valid_bands: list[dict[str, Any]] = []
        score_counter: Counter[str] = Counter()
        for item in items:
            start_dt = to_local_datetime(item.get("startDateTime"), timezone_name)
            end_dt = to_local_datetime(item.get("endDateTime"), timezone_name)
            if start_dt is None or end_dt is None:
                continue
            if not (sunrise_dt <= start_dt < midnight_dt):
                continue
            clipped_end = min(end_dt, midnight_dt)
            scores = item["scores"]
            band_overall = overall_score(scores)
            band = {
                "start": to_short_time(item.get("startDateTime")),
                "end": clipped_end.strftime("%H:%M"),
                "startDateTime": item.get("startDateTime"),
                "endDateTime": clipped_end.isoformat(),
                "primaryState": item.get("primaryState"),
                "riskLevel": item.get("riskLevel"),
                "score": band_overall,
                "categoryScores": {
                    "overall": band_overall,
                    **{key: scores.get(key) for key in SCORE_HEADERS},
                },
            }
            valid_bands.append(band)
            score_counter[band["primaryState"] or "Neutral / Routine"] += 1

        day_sample = next(
            (
                item
                for item in items
                if (to_local_datetime(item.get("startDateTime"), timezone_name) or datetime.min.replace(tzinfo=sunrise_dt.tzinfo))
                >= sunrise_dt
            ),
            sample,
        )

        scored_bands = [
            band
            for band in valid_bands
            if band["categoryScores"].get("overall") is not None
        ]
        best_band = max(scored_bands, key=lambda band: band["categoryScores"]["overall"]) if scored_bands else None
        avoid_scores = [
            band["categoryScores"]["avoid"]
            for band in valid_bands
            if band["categoryScores"].get("avoid") is not None
        ]
        avoid_dominant = bool(avoid_scores) and sum(1 for score in avoid_scores if score >= 80) > len(avoid_scores) / 2
        best_score = best_band["categoryScores"]["overall"] if best_band else None
        if best_score is None:
            day_quality = "Data missing"
        elif avoid_dominant or best_score < 35:
            day_quality = "Avoid"
        elif best_score >= 85:
            day_quality = "Excellent"
        elif best_score >= 70:
            day_quality = "Good"
        elif best_score >= 50:
            day_quality = "Normal"
        else:
            day_quality = "Weak"

        day_summaries.append(
            {
                "date": day_key,
                "day": day_sample.get("day"),
                "sunrise": sunrise_dt.isoformat(),
                "midnight": midnight_dt.isoformat(),
                "mainTithi": day_sample.get("tithi"),
                "mainNakshatra": day_sample.get("moonNakshatra"),
                "bestWindowStart": best_band["start"] if best_band else None,
                "bestWindowEnd": best_band["end"] if best_band else None,
                "bestState": (best_band["primaryState"] if best_band and best_band["primaryState"] else "Data missing"),
                "bestScore": best_score,
                "dayQuality": day_quality,
                "bands": valid_bands,
            }
        )

    first_window = windows[0]
    config_payload = {
        "generatedAt": datetime.now(ZoneInfo(first_window.get("eventTimezone") or "Asia/Kolkata")).isoformat(),
        "sourceWorkbook": display_path(source_workbook),
        "eventLocationName": first_present(windows, "eventLocationName") or normalize_text(config_lookup.get("EventLocationName")),
        "eventLatitude": first_present(windows, "eventLatitude") or to_number(config_lookup.get("EventLatitude")),
        "eventLongitude": first_present(windows, "eventLongitude") or to_number(config_lookup.get("EventLongitude")),
        "eventTimezone": first_present(windows, "eventTimezone") or normalize_text(config_lookup.get("EventTimezone")) or first_window.get("timezone"),
        "natalMoonSign": first_present(windows, "natalMoonSign") or normalize_text(config_lookup.get("NatalMoonSign")),
        "natalNakshatra": first_present(windows, "natalNakshatra") or normalize_text(config_lookup.get("NatalNakshatra")),
        "natalLagna": first_present(windows, "natalLagna") or normalize_text(config_lookup.get("NatalLagna")),
        "availableCategories": AVAILABLE_CATEGORIES,
    }

    exported_dates = sorted(day_groups)
    blocked_error = None
    if score_columns_found and (primary_state_present_count == 0 or all_scores_present_count == 0):
        blocked_error = (
            "V05 formula results are still missing after recalculation. Open the workbook in Excel, "
            "press Ctrl+Alt+Shift+F9, save, close, and rerun exporter."
        )

    if missing_values:
        warnings.append(
            f"{missing_values} rows had missing computed V05 values. If this is unexpected, open the workbook in Excel and recalculate before export."
        )
    if formula_cached_values_missing:
        warnings.append(
            "Formula cached values are missing. Open the V05 workbook in Excel, allow calculation, save, then rerun exporter."
        )

    if blocked_error is not None:
        return ExportStats(
            source_workbook=source_workbook,
            raw_rows_read=raw_rows_read,
            windows_count=len(windows),
            day_count=len(day_summaries),
            first_exported_date=exported_dates[0] if exported_dates else None,
            last_exported_date=exported_dates[-1] if exported_dates else None,
            unique_date_count=len(exported_dates),
            first_sample_dates=exported_dates[:5],
            last_sample_dates=exported_dates[-5:],
            missing_value_count=missing_values,
            missing_score_count=missing_score_count,
            score_columns_found=score_columns_found,
            formula_cached_values_missing=formula_cached_values_missing,
            detected_headers=detected_headers,
            missing_v05_headers=missing_v05_headers,
            debug_samples=debug_samples,
            primary_state_present_count=primary_state_present_count,
            all_scores_present_count=all_scores_present_count,
            excel_recalculation_ran=excel_recalculation_ran,
            excel_recalculation_warning=excel_recalculation_warning,
            blocked_error=blocked_error,
            warnings=warnings,
        )

    WINDOWS_JSON.write_text(json.dumps(windows, indent=2, ensure_ascii=False), encoding="utf-8")
    DAY_SUMMARY_JSON.write_text(json.dumps(day_summaries, indent=2, ensure_ascii=False), encoding="utf-8")
    CONFIG_JSON.write_text(json.dumps(config_payload, indent=2, ensure_ascii=False), encoding="utf-8")
    public_config_payload = {key: value for key, value in config_payload.items() if key != "sourceWorkbook"}
    muhurat_data_payload = {
        "updated_at": public_config_payload.get("generatedAt") or datetime.now(
            ZoneInfo(public_config_payload.get("eventTimezone") or "Asia/Kolkata")
        ).isoformat(),
        "config": public_config_payload,
        "windows": windows,
        "day_summaries": day_summaries,
    }
    MUHURAT_DATA_JSON.write_text(json.dumps(muhurat_data_payload, indent=2, ensure_ascii=False), encoding="utf-8")

    return ExportStats(
        source_workbook=source_workbook,
        raw_rows_read=raw_rows_read,
        windows_count=len(windows),
        day_count=len(day_summaries),
        first_exported_date=exported_dates[0] if exported_dates else None,
        last_exported_date=exported_dates[-1] if exported_dates else None,
        unique_date_count=len(exported_dates),
        first_sample_dates=exported_dates[:5],
        last_sample_dates=exported_dates[-5:],
        missing_value_count=missing_values,
        missing_score_count=missing_score_count,
        score_columns_found=score_columns_found,
        formula_cached_values_missing=formula_cached_values_missing,
        detected_headers=detected_headers,
        missing_v05_headers=missing_v05_headers,
        debug_samples=debug_samples,
        primary_state_present_count=primary_state_present_count,
        all_scores_present_count=all_scores_present_count,
        excel_recalculation_ran=excel_recalculation_ran,
        excel_recalculation_warning=excel_recalculation_warning,
        blocked_error=None,
        warnings=warnings,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Export V06 dashboard JSON from a V05 workbook.")
    parser.add_argument(
        "--source",
        help="Optional source workbook path. Defaults to data_source/MuhuratFinder_V05_ParentStateEngine.xlsx.",
    )
    parser.add_argument(
        "--recalculate",
        dest="recalculate",
        action="store_true",
        default=True,
        help="Force Excel COM recalculation before export. This is the default on Windows.",
    )
    parser.add_argument(
        "--no-recalculate",
        dest="recalculate",
        action="store_false",
        help="Skip Excel COM recalculation before reading cached formula values.",
    )
    args = parser.parse_args()

    source_workbook = choose_source_workbook(args.source)
    stats = export_windows_and_summary(source_workbook, should_recalculate=args.recalculate)
    print(f"SOURCE_WORKBOOK={display_path(stats.source_workbook)}")
    print(f"DETECTED_EPHEMERIS_RAW_HEADERS={', '.join(stats.detected_headers)}")
    print(f"V05_SCORE_COLUMNS_FOUND={stats.score_columns_found}")
    if stats.missing_v05_headers:
        print(f"WARNING=Missing V05 columns: {', '.join(stats.missing_v05_headers)}")
    print(f"EPHEMERIS_RAW_ROWS_READ={stats.raw_rows_read}")
    print(f"WINDOWS_EXPORTED={stats.windows_count}")
    print(f"FIRST_EXPORTED_DATE={stats.first_exported_date}")
    print(f"LAST_EXPORTED_DATE={stats.last_exported_date}")
    print(f"UNIQUE_DATES_IN_WINDOWS={stats.unique_date_count}")
    print(f"DAYS_SUMMARIZED={stats.day_count}")
    print(f"FIRST_5_DATES={', '.join(stats.first_sample_dates)}")
    print(f"LAST_5_DATES={', '.join(stats.last_sample_dates)}")
    print(f"PRIMARY_STATE_PRESENT_ROWS={stats.primary_state_present_count}")
    print(f"ALL_SCORE_VALUES_PRESENT_ROWS={stats.all_scores_present_count}")
    print(f"MISSING_VALUE_ROWS={stats.missing_value_count}")
    print(f"MISSING_SCORE_VALUES={stats.missing_score_count}")
    print(f"FORMULA_CACHED_VALUES_MISSING={stats.formula_cached_values_missing}")
    print(f"EXCEL_COM_RECALCULATION_RAN={stats.excel_recalculation_ran}")
    if stats.excel_recalculation_warning:
        print(f"WARNING={stats.excel_recalculation_warning}")
    for index, sample in enumerate(stats.debug_samples, start=1):
        print(f"DEBUG_SAMPLE_{index}={json.dumps(sample, ensure_ascii=False)}")
    for warning in stats.warnings:
        print(f"WARNING={warning}")
    if stats.blocked_error:
        raise SystemExit(stats.blocked_error)
    print(f"WINDOWS_JSON={display_path(WINDOWS_JSON)}")
    print(f"DAY_SUMMARY_JSON={display_path(DAY_SUMMARY_JSON)}")
    print(f"CONFIG_JSON={display_path(CONFIG_JSON)}")
    print(f"MUHURAT_DATA_JSON={display_path(MUHURAT_DATA_JSON)}")


if __name__ == "__main__":
    main()
